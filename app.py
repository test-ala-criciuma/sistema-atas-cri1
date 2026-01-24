import os
import io
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session, jsonify
from flask_socketio import SocketIO, join_room, leave_room, emit
from functools import wraps
import json
from datetime import datetime, timedelta
import calendar
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors
import models as dbHandler
from functions.pdf_exporters import exportar_pdf_bytes, exportar_sacramental_bytes
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import traceback
from collections import OrderedDict
import secrets

# Excel export
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


app = Flask(__name__)

# Configuração do SocketIO para produção 
try:
    import eventlet
    socketio = SocketIO(app, 
                       cors_allowed_origins="*",
                       async_mode='eventlet')
except ImportError:
    socketio = SocketIO(app, 
                       cors_allowed_origins="*",
                       async_mode='threading')

#Secret key para RENDER
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-123')

# Caminho do DB configurável (permite usar Persistent Disk no Railway)
DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'database', 'atas.db'))

def ensure_db_dir():
    dirpath = os.path.dirname(DB_PATH)
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)


# Conecta BD SQLite
def get_db():
    ensure_db_dir()
    conn = sqlite3.connect(DB_PATH, timeout=5)
    conn.row_factory = sqlite3.Row
    try:
        # Habilita WAL para melhor leitura/escrita simultanea quando suportado
        conn.execute('PRAGMA journal_mode=WAL;')
    except Exception:
        # Em alguns ambientes o PRAGMA pode não ser suportado ou falhar; ignoramos
        pass
    return conn

# Inicializa BD
def init_db():
    """Inicializa o banco a partir do schema se o arquivo DB não existir, ou se
    o arquivo existir mas estiver vazio/sem tabelas criadas (caso comum em cópias
    de arquivo ou DB corrompido). Evita reexecutar o schema em um BD já populado
    (previne erros como "duplicate column name")."""
    db_path = DB_PATH
    ensure_db_dir()

    # Caso 1: arquivo não existe → criar e aplicar schema
    if not os.path.exists(db_path):
        should_init = True
    else:
        # Arquivo existe — primeiro verificar se é um arquivo SQLite válido (assinatura)
        try:
            with open(db_path, 'rb') as fb:
                sig = fb.read(16)
            if not sig.startswith(b'SQLite format 3\x00'):
                # Arquivo inválido (possivelmente HTML/corrompido). Mover para backups e forçar init
                backup_dir = os.path.join(os.path.dirname(db_path), 'backups')
                os.makedirs(backup_dir, exist_ok=True)
                ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                corrupt_path = os.path.join(backup_dir, f'atas.db.corrupt.{ts}')
                os.replace(db_path, corrupt_path)
                print(f"Arquivo DB inválido movido para backup: {corrupt_path}. Será criado um novo DB.")
                should_init = True
        except Exception as e:
            # Se não conseguimos ler o arquivo, marcar para inicializar e tentar corrigir
            print(f"Erro ao verificar assinatura do DB: {e}. Tentando aplicar schema.")
            should_init = True

        # Se o arquivo parecia ser SQLite, verificar se tem as tabelas mínimas
        if not locals().get('should_init'):
            try:
                conn = get_db()
                tbls = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
                conn.close()
                # Se as tabelas mínimas estiverem presentes, NÃO inicializamos; caso contrário, aplicamos schema
                should_init = not ('atas' in tbls and 'sacramental' in tbls and 'users' in tbls)
                if should_init:
                    print("Banco existe mas sem tabelas completas — aplicando schema_inicial.sql")
            except Exception as e:
                # Problema ao inspecionar — tentar inicializar para corrigir
                print(f"Erro ao inspecionar DB existente: {e}. Tentando aplicar schema.")
                should_init = True

    if not should_init:
        return

    with app.app_context():
        conn = get_db()
        try:
            with open('database/schema_inicial.sql', 'r', encoding='utf-8') as f:
                sql_script = f.read()
            conn.executescript(sql_script)
            conn.commit()
            conn.close()
            print("Banco de dados inicializado com sucesso.")
        except Exception as e:
            print(f"Erro ao inicializar banco: {e}")

# Garantir colunas adicionadas para compatibilidade com versões antigas
def ensure_sacramental_columns():
    conn = get_db()
    try:
        cur = conn.execute("PRAGMA table_info(sacramental)").fetchall()
    except Exception:
        # Tabela não existe — nada a migrar aqui
        conn.close()
        return

    # Se o PRAGMA retornar vazio, a tabela provavelmente não existe no banco atual
    if not cur:
        print("Tabela 'sacramental' não encontrada — pulando verificação de colunas.")
        conn.close()
        return

    existing = [c['name'] for c in cur]

    to_add = {
        'discursante_1': 'TEXT',
        'discursante_2': 'TEXT',
        'outros': 'TEXT',
        'tema_1': 'TEXT',
        'tema_2': 'TEXT',
        'tema_ultimo': 'TEXT',
        'obs_1': 'TEXT',
        'obs_2': 'TEXT',
        'obs_ultimo': 'TEXT'
    }

    for col, coltype in to_add.items():
        if col not in existing:
            try:
                conn.execute(f"ALTER TABLE sacramental ADD COLUMN {col} {coltype}")
                print(f"Coluna {col} adicionada em sacramental")
            except Exception as e:
                print(f"Não foi possível adicionar coluna {col}: {e}")
    conn.commit()
    conn.close()

# Executar checagem de colunas no startup (silencioso)
try:
    ensure_sacramental_columns()
except Exception:
    pass

# Garantir que o DB exista e o schema seja aplicado em qualquer ambiente de execução
# (útil quando o app é iniciado por Gunicorn / Railway — evita dependência do bloco __main__)
try:
    init_db()
except Exception as e:
    print(f"Erro ao inicializar DB no startup: {e}")

# Mensagem Autenticação no Login
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Por favor, faça login para acessar esta página.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# app.py

# Autenticação Login
def authenticate_user(username, password):
    conn = get_db()
    # 1. Busca o usuário APENAS pelo username (NUNCA pela senha)
    user = conn.execute(
        "SELECT * FROM users WHERE username = ?", 
        (username,)
    ).fetchone()
    conn.close()
    
    # 2. Se o usuário existir, verifica a senha contra o hash armazenado
    if user and check_password_hash(user['password'], password):
        return user
    return None # Retorna None se o usuário não for encontrado ou a senha não bater

# ==================================================================
# Rotas principais do sistema de atas
# ==================================================================

# Aba de discursantes recentes na criação de atas sacramentais
def get_discursantes_recentes():
    """Busca discursantes dos últimos 3 meses agrupados por data."""
    conn = get_db()
    tres_meses_atras = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
    
    # Seleciona os dados
    # Compatibilidade: só referenciar a coluna antiga 'discursantes' se ela existir no esquema
    has_discursantes_col = any(c['name'] == 'discursantes' for c in conn.execute("PRAGMA table_info(sacramental)").fetchall())
    if has_discursantes_col:
        registros = conn.execute("""
            SELECT s.discursantes, s.discursante_1, s.discursante_2, s.ultimo_discursante, a.data
            FROM sacramental s 
            JOIN atas a ON s.ata_id = a.id 
            WHERE a.data >= ? AND a.tipo = 'sacramental' AND a.ala_id = ?
            ORDER BY a.data DESC
        """, (tres_meses_atras, session['user_id'])).fetchall()
    else:
        registros = conn.execute("""
            SELECT s.discursante_1, s.discursante_2, s.ultimo_discursante, a.data
            FROM sacramental s 
            JOIN atas a ON s.ata_id = a.id 
            WHERE a.data >= ? AND a.tipo = 'sacramental' AND a.ala_id = ?
            ORDER BY a.data DESC
        """, (tres_meses_atras, session['user_id'])).fetchall()
    
    agrupado_por_data = []
    
    for row in registros:
        # converter sqlite3.Row para dict para permitir row.get(...) sem erro
        row = dict(row)
        lista_nomes = []
        # 1. Preferir colunas individuais
        if row.get('discursante_1') and row['discursante_1'].strip():
            lista_nomes.append(row['discursante_1'].strip())
        if row.get('discursante_2') and row['discursante_2'].strip():
            lista_nomes.append(row['discursante_2'].strip())
        # 2. Fallback para campo JSON antigo
        if not lista_nomes and row.get('discursantes'):
            try:
                nomes_json = json.loads(row['discursantes'])
                lista_nomes.extend([n.strip() for n in nomes_json if n and n.strip()])
            except: pass
        # 3. Pega o último discursante
        if row.get('ultimo_discursante') and row['ultimo_discursante'].strip():
            lista_nomes.append(row['ultimo_discursante'].strip())
        # 4. Se houver nomes, adiciona ao grupo daquela data
        if lista_nomes:
            data_fmt = datetime.strptime(row['data'], "%Y-%m-%d").strftime("%d/%m/%Y")
            agrupado_por_data.append({
                'data': data_fmt,
                'nomes': lista_nomes
            })
    
    return agrupado_por_data[:20] # Retorna as últimas 20 reuniões

# Próxima reunião sacramental automática na página inicial
def get_proxima_reuniao_sacramental():
    """Encontra a data da próxima reunião sacramental e verifica se já existe ata
    para a ala do usuário logado. Sempre retorna um dicionário com a data; se
    existir uma ata para a ala atual, inclui 'ata_existente': True e o 'id'."""
    hoje = datetime.now().date()

    # Encontrar próximo domingo
    dias_para_domingo = (6 - hoje.weekday()) % 7
    if dias_para_domingo == 0:  # Se hoje é domingo
        proximo_domingo = hoje
    else:
        proximo_domingo = hoje + timedelta(days=dias_para_domingo)

    # Formatar data em português
    data_formatada = proximo_domingo.strftime("%d/%m/%Y")

    # Verificar se já existe ata para esta data E para a ala do usuário logado
    conn = get_db()
    try:
        ala_id = session.get('user_id')
        ata_row = conn.execute(
            "SELECT * FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?",
            (proximo_domingo.strftime("%Y-%m-%d"), ala_id)
        ).fetchone()
    finally:
        conn.close()

    if ata_row:
        return {
            'data': proximo_domingo.strftime("%Y-%m-%d"),
            'data_formatada': data_formatada,
            'ata_existente': True,
            'id': ata_row['id']
        }
    else:
        return {
            'data': proximo_domingo.strftime("%Y-%m-%d"),
            'data_formatada': data_formatada,
            'ata_existente': False
        }
def get_temas_recentes():
    """Busca temas dos últimos 3 meses"""
    conn = get_db()
    
    # Data de 90 dias atrás
    tres_meses_atras = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
    
    temas_recentes = conn.execute("""
        SELECT DISTINCT s.tema, a.data 
        FROM sacramental s 
        JOIN atas a ON s.ata_id = a.id 
        WHERE date(a.data) >= date(?) 
          AND a.tipo = 'sacramental' 
          AND a.ala_id = ? 
          AND s.tema IS NOT NULL 
          AND TRIM(s.tema) <> ''
        ORDER BY a.data DESC
        LIMIT 10
    """, (tres_meses_atras, session['user_id'])).fetchall()
    
    temas_formatados = []
    for tema in temas_recentes:
        if tema['tema']:
            data_obj = datetime.strptime(tema['data'], "%Y-%m-%d")
            data_formatada = data_obj.strftime("%d/%m/%Y")
            temas_formatados.append({
                'tema': tema['tema'],
                'data': data_formatada
            })
    
    conn.close()
    return temas_formatados[:10]

def get_hinos_recentes():
    """Busca hinos tocados nos últimos 2 meses, agrupados por data e ordem litúrgica."""
    conn = get_db()
    
    dois_meses_atras = (datetime.now() - timedelta(days=60)).strftime("%Y-%m-%d")
    
    hinos_recentes_raw = conn.execute("""
        SELECT a.data, s.hinos, s.hino_sacramental, s.hino_intermediario
        FROM sacramental s
        JOIN atas a ON s.ata_id = a.id
        WHERE date(a.data) >= date(?)
          AND a.tipo = 'sacramental'
          AND a.ala_id = ?
        ORDER BY a.data DESC
        LIMIT 10
    """, (dois_meses_atras, session['user_id'])).fetchall()

    hinos_por_data = {}

    for row in hinos_recentes_raw:
        data_obj = datetime.strptime(row['data'], "%Y-%m-%d")
        data_formatada = data_obj.strftime("%d/%m/%Y")

        hinos_lista = []
        
        # 1. Extrair os hinos do JSON [abertura, encerramento] primeiro
        h_abertura = ""
        h_encerramento = ""
        try:
            hinos_json = json.loads(row['hinos'] or '[]')
            if len(hinos_json) > 0: h_abertura = hinos_json[0]
            if len(hinos_json) > 1: h_encerramento = hinos_json[1]
        except json.JSONDecodeError: 
            pass

        # 2. Montar a lista na ORDEM CORRETA
        
        # Abertura
        if h_abertura and h_abertura.strip(): 
            hinos_lista.append({'tipo': 'Abertura', 'nome': h_abertura.strip()})
        
        # Sacramental
        if row['hino_sacramental'] and row['hino_sacramental'].strip(): 
            hinos_lista.append({'tipo': 'Sacramental', 'nome': row['hino_sacramental'].strip()})
        
        # Intermediário
        if row['hino_intermediario'] and row['hino_intermediario'].strip(): 
            hinos_lista.append({'tipo': 'Intermediário', 'nome': row['hino_intermediario'].strip()})
            
        # Encerramento
        if h_encerramento and h_encerramento.strip(): 
            hinos_lista.append({'tipo': 'Encerramento', 'nome': h_encerramento.strip()})

        if hinos_lista and data_formatada not in hinos_por_data:
            hinos_por_data[data_formatada] = {'data': data_formatada, 'hinos': hinos_lista}
            
    return list(hinos_por_data.values())[:10]

# Configuração do Rate Limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    storage_uri="memory://",  # Use memória para desenvolvimento. Em produção, use Redis ou Memcached.
    default_limits=["200 per day", "50 per hour"]
)

# Rota de Login de Usuário
@app.route('/', methods=['GET', 'POST'])
@limiter.limit("5 per minute", methods=["POST"]) # error="Muitas tentativas de login. Tente novamente em um minuto.")
def login():
    # If user is already logged in, redirect to index
    if session.get('logged_in'):
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        if not username or not password:
            flash('Por favor, preencha todos os campos.', 'error')
            return render_template('login.html')
        
        user = authenticate_user(username, password)
        
        if user:
            session['logged_in'] = True
            session['username'] = user['username']
            session['user_id'] = user['id']
            flash(f'Login realizado com sucesso! Bem-vindo, {user["username"]}.', 'success')
            return redirect(url_for('index'))
        else:
            flash('Credenciais inválidas. Por favor, tente novamente.', 'error')
    
    return render_template('login.html')

# Rota de Logout de Usuário
@app.route('/logout')
def logout():
    session.clear()
    flash('Você saiu do sistema.', 'success')
    return redirect(url_for('login'))

# ==================================================================
# Rotas de Configurações
# ==================================================================

# Rota para configurações com clonagem automática de templates padrão
@app.route("/configuracoes")
@login_required
def configuracoes():
    conn = get_db()
    ala_id = session['user_id']

    # 1. Buscar templates da ala logada
    templates_row = conn.execute("SELECT * FROM templates WHERE ala_id = ?", (ala_id,)).fetchall()
    
    # 2. LÓGICA DE CLONAGEM: Se não houver templates para esta ala, copia os padrões (ala_id = 0)
    if not templates_row:
        modelos_mestres = conn.execute("SELECT * FROM templates WHERE ala_id = 0").fetchall()
        for modelo in modelos_mestres:
            conn.execute("""
                INSERT INTO templates (
                    ala_id, tipo_template, nome, boas_vindas, desobrigacoes, apoios, 
                    confirmacoes_batismo, apoio_membro_novo, bencao_crianca, 
                    sacramento, mensagens, live, encerramento
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                ala_id, modelo['tipo_template'], modelo['nome'], modelo['boas_vindas'], 
                modelo['desobrigacoes'], modelo['apoios'], modelo['confirmacoes_batismo'], 
                modelo['apoio_membro_novo'], modelo['bencao_crianca'], modelo['sacramento'], 
                modelo['mensagens'], modelo['live'], modelo['encerramento']
            ))
        conn.commit()
        # Refaz a busca agora com os templates clonados
        templates_row = conn.execute("SELECT * FROM templates WHERE ala_id = ?", (ala_id,)).fetchall()

    templates = [dict(t) for t in templates_row]

    # 3. Buscar informações da unidade (Sua lógica original preservada)
    unidade_row = conn.execute(
        "SELECT * FROM unidades WHERE ala_id = ?",
        (ala_id,)
    ).fetchone()

    if unidade_row:
        unidade = dict(unidade_row)
        primeiro = unidade.get('primeiro_conselheiro') or ''
        segundo = unidade.get('segundo_conselheiro') or ''

        if not primeiro and not segundo:
            cons_raw = unidade.get('conselheiros') or ''
            if cons_raw:
                try:
                    parsed = json.loads(cons_raw)
                    if isinstance(parsed, list):
                        primeiro = parsed[0] if len(parsed) > 0 else ''
                        segundo = parsed[1] if len(parsed) > 1 else ''
                except Exception:
                    if '|' in cons_raw:
                        parts = [p.strip() for p in cons_raw.split('|', 1)]
                        primeiro = parts[0]
                        segundo = parts[1] if len(parts) > 1 else ''
                    elif '\n' in cons_raw:
                        parts = [p.strip() for p in cons_raw.split('\n', 1)]
                        primeiro = parts[0]
                        segundo = parts[1] if len(parts) > 1 else ''
                    else:
                        primeiro = cons_raw.strip()

        unidade['primeiro_conselheiro'] = primeiro
        unidade['segundo_conselheiro'] = segundo
    else:
        unidade = {}

    # 4. Buscar estatísticas (Sua lógica original preservada)
    total_atas = conn.execute(
        "SELECT COUNT(*) FROM atas WHERE ala_id = ?",
        (ala_id,)
    ).fetchone()[0]

    atas_sacramentais = conn.execute(
        "SELECT COUNT(*) FROM atas WHERE ala_id = ? AND tipo = 'sacramental'",
        (ala_id,)
    ).fetchone()[0]

    atas_batismo = conn.execute(
        "SELECT COUNT(*) FROM atas WHERE ala_id = ? AND tipo = 'batismo'",
        (ala_id,)
    ).fetchone()[0]

    mes_atual = datetime.now().strftime("%Y-%m")
    atas_mes = conn.execute(
        "SELECT COUNT(*) FROM atas WHERE ala_id = ? AND strftime('%Y-%m', data) = ?",
        (ala_id, mes_atual)
    ).fetchone()[0]

    conn.close()

    return render_template(
        "configuracoes.html",
        templates=templates,
        unidade=unidade,
        total_atas=total_atas,
        atas_sacramentais=atas_sacramentais,
        atas_batismo=atas_batismo,
        atas_mes=atas_mes
    )

# Rota para salvar configurações da ala
@app.route("/configuracoes/ala/salvar", methods=["POST"])
@login_required
def salvar_configuracoes_ala():
    conn = get_db()

    nome_ala = request.form.get("nome_ala")
    bispo = request.form.get("bispo")
    primeiro_conselheiro = request.form.get("primeiro_conselheiro")
    segundo_conselheiro = request.form.get("segundo_conselheiro")
    recepcionista = request.form.get("recepcionista")
    pianista = request.form.get("pianista")
    regente_musica = request.form.get("regente_musica")
    horario = request.form.get("horario")
    
    # Verificar se já existe registro para esta ala
    unidade_existente = conn.execute(
        "SELECT * FROM unidades WHERE ala_id = ?",
        (session['user_id'],)
    ).fetchone()

    if unidade_existente:
        # Atualizar - não tocar em estaca_id para evitar inconsistências
        conn.execute("""
            UPDATE unidades
            SET nome = ?, bispo = ?, primeiro_conselheiro = ?, segundo_conselheiro = ?, horario = ?, recepcionista = ?, pianista = ?, regente_musica = ?
            WHERE ala_id = ?
        """, (nome_ala, bispo, primeiro_conselheiro, segundo_conselheiro, horario, recepcionista, pianista, regente_musica, session['user_id']))
    else:
        # Inserir - estaca_id usará valor default definido no schema (DEFAULT 1)
        conn.execute("""
            INSERT INTO unidades (ala_id, nome, bispo, primeiro_conselheiro, segundo_conselheiro, horario, recepcionista, pianista, regente_musica)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (session['user_id'], nome_ala, bispo, primeiro_conselheiro, segundo_conselheiro, horario, recepcionista, pianista, regente_musica))

    conn.commit()
    conn.close()

    flash("Configurações da ala salvas com sucesso!", "success")
    return redirect(url_for("configuracoes"))

# Rota para editar template
@app.route("/configuracoes/template/<int:template_id>")
@login_required
def editar_template(template_id):
    conn = get_db()
    template = conn.execute(
        "SELECT * FROM templates WHERE id = ?", 
        (template_id,)
    ).fetchone()
    
    if template:
        template = dict(template)
        conn.close()
        return render_template("_editar_template.html", template=template)
    else:
        conn.close()
        return "Template não encontrado", 404

# Rota para visualizar template (somente leitura)
@app.route("/configuracoes/template/<int:template_id>/visualizar")
@login_required
def visualizar_template(template_id):
    conn = get_db()
    template = conn.execute(
        "SELECT * FROM templates WHERE id = ?",
        (template_id,)
    ).fetchone()

    if template:
        template = dict(template)
        conn.close()
        return render_template("_visualizar_template.html", template=template)
    else:
        conn.close()
        return "Template não encontrado", 404

# Rota para salvar template
@app.route("/configuracoes/template/<int:template_id>/salvar", methods=["POST"])
@login_required
def salvar_template(template_id):
    conn = get_db()
    try:
        # Mapeamento exato com o seu novo SCHEMA do SQL
        conn.execute("""
            UPDATE templates SET
                nome = ?, boas_vindas = ?, desobrigacoes = ?, apoios = ?, 
                confirmacoes_batismo = ?, apoio_membro_novo = ?, bencao_crianca = ?,
                sacramento = ?, mensagens = ?, live = ?, encerramento = ?
            WHERE id = ? AND ala_id = ?
        """, (
            request.form.get('nome'), request.form.get('boas_vindas'),
            request.form.get('desobrigacoes'), request.form.get('apoios'),
            request.form.get('confirmacoes_batismo'), request.form.get('apoio_membro_novo'),
            request.form.get('bencao_crianca'), request.form.get('sacramento'),
            request.form.get('mensagens'), request.form.get('live'),
            request.form.get('encerramento'), template_id, session['user_id']
        ))
        
        conn.commit()
        # Retornar JSON se AJAX, para permitir atualização fluida do frontend
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            updated = conn.execute("SELECT * FROM templates WHERE id = ?", (template_id,)).fetchone()
            conn.close()
            return jsonify({ 'success': True, 'message': 'Template atualizado com sucesso!', 'template': dict(updated) })

        flash("Template atualizado com sucesso!", "success")
    except Exception as e:
        flash(f"Erro ao salvar: {e}", "error")
    finally:
        conn.close()
    return redirect(url_for("configuracoes"))

# Rota para criar novo template
@app.route("/configuracoes/template/criar", methods=["POST"])
@login_required
def criar_template():
    conn = get_db()
    ala_id = session.get('user_id')
    
    try:
        nome = request.form.get('nome')
        tipo_template = request.form.get('tipo_template') # 1=Sacramental, 2=Batismo
        
        # 1. VERIFICAÇÃO DE DUPLICIDADE: 
        # Busca se já existe um template desse TIPO para essa ALA
        existente = conn.execute(
            "SELECT id FROM templates WHERE tipo_template = ? AND ala_id = ?", 
            (tipo_template, ala_id)
        ).fetchone()

        if existente:
            # Se já existe, responder apropriadamente
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                conn.close()
                return jsonify({ 'success': False, 'message': 'Já existe um template para este tipo. Por favor, edite o existente.' }), 400
            # O ideal é que o usuário use a rota de SALVAR para editar.
            flash("Já existe um template para este tipo. Por favor, edite o existente.", "warning")
            return redirect(url_for("configuracoes"))

        # 2. INSERÇÃO (Caso seja realmente novo)
        # Ler todos os campos enviados pelo formulário
        boas_vindas = request.form.get('boas_vindas') or ''
        desobrigacoes = request.form.get('desobrigacoes') or ''
        apoios = request.form.get('apoios') or ''
        confirmacoes_batismo = request.form.get('confirmacoes_batismo') or ''
        apoio_membro_novo = request.form.get('apoio_membro_novo') or ''
        bencao_crianca = request.form.get('bencao_crianca') or ''
        sacramento = request.form.get('sacramento') or ''
        mensagens = request.form.get('mensagens') or ''
        live = request.form.get('live') or ''
        encerramento = request.form.get('encerramento') or ''

        # Use cursor to get lastrowid so we can return created template for AJAX
        cur = conn.execute("""
            INSERT INTO templates (
                tipo_template, ala_id, nome, boas_vindas, desobrigacoes, apoios, 
                confirmacoes_batismo, apoio_membro_novo, bencao_crianca, 
                sacramento, mensagens, live, encerramento
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            tipo_template, ala_id, nome, boas_vindas, desobrigacoes, apoios,
            confirmacoes_batismo, apoio_membro_novo, bencao_crianca,
            sacramento, mensagens, live, encerramento
        ))
        conn.commit()
        new_id = cur.lastrowid

        # Se requisição via AJAX, retornar JSON com os dados do template
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            new_tpl = conn.execute("SELECT * FROM templates WHERE id = ?", (new_id,)).fetchone()
            conn.close()
            tpl = dict(new_tpl)
            return jsonify({ 'success': True, 'message': 'Novo template criado com sucesso!', 'template': tpl })

        flash("Novo template criado com sucesso!", "success")
    except Exception as e:
        print(f"Erro: {e}")
        flash("Erro ao criar template", "error")
    finally:
        conn.close()
    return redirect(url_for("configuracoes"))
   
# Rota para exportar dados em Excel (apenas ATAs)
@app.route('/configuracoes/exportar')
@login_required
def exportar_dados():
    conn = get_db()
    try:
        ala_id = session.get('user_id')

        # Buscar todas as atas da ala
        atas = conn.execute("SELECT * FROM atas WHERE ala_id = ? ORDER BY data DESC", (ala_id,)).fetchall()

        wb = Workbook()

        # Planilha principal com ATAS e campos detalhados (sacramental + batismo)
        ws = wb.active
        ws.title = 'Atas'

        headers = [
            'id','tipo','data','status','ala_id',
            # Sacramental fields
            'tema','presidido','dirigido','pianista','regente_musica','anuncios','hinos','hino_sacramental','hino_intermediario','oracoes','discursantes','recepcionistas','reconhecemos_presenca','desobrigacoes','apoios','confirmacoes_batismo','apoio_membros','bencao_criancas','ultimo_discursante','id_tipo',
            # Batismo fields
            'dedicado','batizados','testemunha1','testemunha2'
        ]
        ws.append(headers)

        for a in atas:
            # Tentar buscar dados sacramentais e de batismo relacionados
            s = conn.execute("SELECT * FROM sacramental WHERE ata_id = ?", (a['id'],)).fetchone()
            b = conn.execute("SELECT * FROM batismo WHERE ata_id = ?", (a['id'],)).fetchone()

            row = [
                a['id'], a['tipo'], a['data'], a['status'], a['ala_id'],
                # Sacramental values (se houver)
                (s['tema'] if s and s['tema'] is not None else ''),
                (s['presidido'] if s and s['presidido'] is not None else ''),
                (s['dirigido'] if s and s['dirigido'] is not None else ''),
                (s['pianista'] if s and s['pianista'] is not None else ''),
                (s['regente_musica'] if s and s['regente_musica'] is not None else ''),
                (s['anuncios'] if s and s['anuncios'] is not None else ''),
                (s['hinos'] if s and s['hinos'] is not None else ''),
                (s['hino_sacramental'] if s and s['hino_sacramental'] is not None else ''),
                (s['hino_intermediario'] if s and s['hino_intermediario'] is not None else ''),
                (s['oracoes'] if s and s['oracoes'] is not None else ''),
                (s['discursantes'] if s and s['discursantes'] is not None else ''),
                (s['recepcionistas'] if s and s['recepcionistas'] is not None else ''),
                (s['reconhecemos_presenca'] if s and s['reconhecemos_presenca'] is not None else ''),
                (s['desobrigacoes'] if s and s['desobrigacoes'] is not None else ''),
                (s['apoios'] if s and s['apoios'] is not None else ''),
                (s['confirmacoes_batismo'] if s and s['confirmacoes_batismo'] is not None else ''),
                (s['apoio_membros'] if s and s['apoio_membros'] is not None else ''),
                (s['bencao_criancas'] if s and s['bencao_criancas'] is not None else ''),
                (s['ultimo_discursante'] if s and s['ultimo_discursante'] is not None else ''),
                (s['id_tipo'] if s and s['id_tipo'] is not None else ''),
                # Batismo values (se houver)
                (b['dedicado'] if b and b['dedicado'] is not None else ''),
                (b['batizados'] if b and b['batizados'] is not None else ''),
                (b['testemunha1'] if b and b['testemunha1'] is not None else ''),
                (b['testemunha2'] if b and b['testemunha2'] is not None else '')
            ]

            ws.append(row)

        # Auto-size columns for clareza
        for i, column_cells in enumerate(ws.columns, 1):
            length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = min(max(length + 2, 10), 60)

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"atas_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(bio,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=filename)
    except Exception as e:
        print(f"Erro exportar xlsx: {e}")
        return ("Erro na exportação", 500)
    finally:
        conn.close()

# Rota para gerar e baixar backup do banco (requer senha ou sessão)
@app.route('/configuracoes/backup', methods=['POST'])
def fazer_backup_db():
    try:
        data = None
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form
        senha = (data.get('password') or '').strip()

        # BACKUP_PASSWORD deve estar configurada no ambiente para permitir downloads sem sessão
        BACKUP_PASSWORD = os.environ.get('BACKUP_PASSWORD')
        if not BACKUP_PASSWORD:
            # Não prosseguir sem uma senha configurada no servidor
            return jsonify({'success': False, 'message': 'Backup password is not configured on the server.'}), 500

        # Autorização: ou usuário logado, ou senha correta via BACKUP_PASSWORD
        authorized = False
        if session.get('logged_in'):
            authorized = True
        else:
            if senha and secrets.compare_digest(senha, BACKUP_PASSWORD):
                authorized = True

        if not authorized:
            # Se não autorizado, retornamos 403 (sem redirecionar à página de login)
            return jsonify({'success': False, 'message': 'Não autorizado'}), 403

        db_path = DB_PATH
        if not os.path.exists(db_path):
            return jsonify({'success': False, 'message': 'Arquivo de banco não encontrado'}), 404

        filename = f"atas_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db"
        return send_file(db_path, mimetype='application/x-sqlite3', as_attachment=True, download_name=filename)
    except Exception as e:
        print(f"Erro ao gerar backup: {e}")
        return jsonify({'success': False, 'message': 'Erro ao gerar backup'}), 500

# Rota para apagar template
@app.route("/configuracoes/template/<int:template_id>/apagar", methods=["POST"])
@login_required
def apagar_template(template_id):
    conn = get_db()
    
    try:
        # Verificar se o template existe
        template = conn.execute(
            "SELECT * FROM templates WHERE id = ?", 
            (template_id,)
        ).fetchone()
        
        if not template:
            return jsonify({
                'success': False,
                'message': 'Template não encontrado'
            }), 404
        
        # Não permitir apagar todos os templates - manter pelo menos um de cada tipo
        templates_restantes = conn.execute(
            "SELECT COUNT(*) FROM templates WHERE tipo_template = ?", 
            (template['tipo_template'],)
        ).fetchone()[0]
        
        if templates_restantes <= 1:
            return jsonify({
                'success': False,
                'message': 'Não é possível apagar o último template deste tipo'
            }, 400)
        
        # Apagar o template
        conn.execute("DELETE FROM templates WHERE id = ?", (template_id,))
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True,
            'message': 'Template apagado com sucesso!'
        })
        
    except Exception as e:
        conn.close()
        print(f"Erro ao apagar template: {e}")
        return jsonify({
            'success': False,
            'message': 'Erro interno ao apagar template'
        }), 500

# ==================================================================
# Rotas de Atas
# ==================================================================

# Página Inicial com lista de atas
@app.route('/index')
@login_required
def index():
    conn = get_db()
    
    # Gerar lista de meses para o seletor EM PORTUGUÊS
    meses = []
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # Nomes dos meses em português
    meses_ptbr = [
        '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]
    
    for month in range(1, 13):
        month_name = meses_ptbr[month]
        month_value = f"{current_year}-{month:02d}"
        meses.append({
            'value': month_value,
            'nome': f"{month_name} {current_year}"
        })
    
    # Formato do mês atual para seleção automática
    mes_atual = datetime.now().strftime("%Y-%m")
    mes_nome = meses_ptbr[datetime.now().month] + " " + str(datetime.now().year)  # CORREÇÃO: Definir mes_nome
    
    # Carregar atas do mês atual da ala do usuário
    atas = conn.execute(
        "SELECT * FROM atas WHERE strftime('%Y-%m', data) = ? AND ala_id = ? ORDER BY data DESC", 
        (mes_atual, session['user_id'])
    ).fetchall()
    
    # Buscar próxima reunião sacramental
    proxima_reuniao = get_proxima_reuniao_sacramental()
    
    return render_template(
        "index.html",
        meses=meses,
        mes_atual=mes_atual,
        mes_nome=mes_nome,  # AGORA ESTÁ DEFINIDA
        atas=atas,
        proxima_reuniao=proxima_reuniao
    )

# Rota para visualizar todas as atas
@app.route("/atas")
@login_required
def listar_todas_atas():
    conn = get_db()
    # Para garantir que o retorno do banco seja acessível por nome de coluna
    conn.row_factory = sqlite3.Row 
    
    # 1. Buscar todas as atas da ala
    atas = conn.execute("""
        SELECT a.*, s.tema 
        FROM atas a 
        LEFT JOIN sacramental s ON a.id = s.ata_id 
        WHERE a.ala_id = ? 
        ORDER BY a.data DESC
    """, (session['user_id'],)).fetchall()
    
    # 2. Lógica idêntica ao criar/editar para Discursantes Recentes
    tres_meses_atras = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")
    
    # Compatibilidade: mantém suporte para o campo JSON antigo 'discursantes' apenas se existir
    has_discursantes_col = any(c['name'] == 'discursantes' for c in conn.execute("PRAGMA table_info(sacramental)").fetchall())
    if has_discursantes_col:
        registros = conn.execute("""
            SELECT s.discursante_1, s.discursante_2, s.discursantes, s.ultimo_discursante, a.data
            FROM sacramental s 
            JOIN atas a ON s.ata_id = a.id 
            WHERE a.data >= ? AND a.tipo = 'sacramental' AND a.ala_id = ?
            ORDER BY a.data DESC
        """, (tres_meses_atras, session['user_id'])).fetchall()
    else:
        registros = conn.execute("""
            SELECT s.discursante_1, s.discursante_2, s.ultimo_discursante, a.data
            FROM sacramental s 
            JOIN atas a ON s.ata_id = a.id 
            WHERE a.data >= ? AND a.tipo = 'sacramental' AND a.ala_id = ?
            ORDER BY a.data DESC
        """, (tres_meses_atras, session['user_id'])).fetchall()
    
    agrupado_por_data = []
    
    for row in registros:
        # converter sqlite3.Row para dict para permitir row.get(...) sem erro
        row = dict(row)
        lista_nomes = []
        # 1. Preferir colunas individuais (compatível com novo esquema)
        if row.get('discursante_1') and row['discursante_1'].strip():
            lista_nomes.append(row['discursante_1'].strip())
        if row.get('discursante_2') and row['discursante_2'].strip():
            lista_nomes.append(row['discursante_2'].strip())
        # 2. Fallback para campo JSON antigo caso colunas individuais estejam vazias
        if not lista_nomes and row.get('discursantes'):
            try:
                nomes_json = json.loads(row['discursantes'])
                lista_nomes.extend([n.strip() for n in nomes_json if n and n.strip()])
            except: pass
        # 3. Pega o último discursante
        if row.get('ultimo_discursante') and row['ultimo_discursante'].strip():
            lista_nomes.append(row['ultimo_discursante'].strip())
        # 4. Se houver nomes, adiciona ao grupo daquela data
        if lista_nomes:
            data_fmt = datetime.strptime(row['data'], "%Y-%m-%d").strftime("%d/%m/%Y")
            agrupado_por_data.append({
                'data': data_fmt,
                'nomes': lista_nomes  # Aqui está a chave que o seu HTML (for nome in grupo.nomes) precisa!
            })

    # 3. Temas Recentes (Formatando a data corretamente)
    temas_db = conn.execute("""
        SELECT s.tema, a.data 
        FROM sacramental s 
        JOIN atas a ON s.ata_id = a.id 
        WHERE a.data >= ? AND a.tipo = 'sacramental' AND a.ala_id = ?
          AND s.tema IS NOT NULL AND TRIM(s.tema) <> ''
        ORDER BY a.data DESC
    """, (tres_meses_atras, session['user_id'])).fetchall()

    temas_formatados = []
    for t in temas_db:
        temas_formatados.append({
            'tema': t['tema'],
            'data': datetime.strptime(t['data'], "%Y-%m-%d").strftime("%d/%m/%Y")
        })
    
    # Build available months list (unique YYYY-MM with label)
    available_months = []
    seen_months = set()
    meses_ptbr = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                  'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    for a in atas:
        ym = a['data'][:7]
        if ym not in seen_months:
            seen_months.add(ym)
            y, m = ym.split('-')
            label = meses_ptbr[int(m)] + ' ' + y
            available_months.append({'value': ym, 'label': label})

    conn.close()
    
    # Passa data da próxima reunião sacramental para exibir no sidebar (opcional)
    proxima_reuniao = get_proxima_reuniao_sacramental()
    
    # As rotas para gerenciar Discursantes e Temas foram movidas para o nível do módulo (definidas abaixo) para evitar problemas de escopo ao construir URLs.

    return render_template(
        "todas_atas.html",
        atas=atas,
        available_months=available_months,
        discursantes_recentes=agrupado_por_data[:20],
        temas_recentes=temas_formatados[:20],
        hinos_recentes=get_hinos_recentes(),
        data=proxima_reuniao['data'] if proxima_reuniao else None
    )

# Rotas para Discursantes e Temas (nível do módulo)
@app.route("/discursantes_temas", methods=["GET"])
@login_required
def discursantes_temas():
    # original implementation below (unchanged)

    hoje = datetime.now().date()

    # Calcular fim como o último dia do mês que está 3 meses à frente para evitar cortar no meio
    months_to_add = 3
    target_month = (hoje.month - 1 + months_to_add) % 12 + 1
    target_year = hoje.year + (hoje.month - 1 + months_to_add) // 12
    last_day = calendar.monthrange(target_year, target_month)[1]
    fim = datetime(year=target_year, month=target_month, day=last_day).date()

    # Próximo domingo (ou hoje se domingo)
    dias_para_domingo = (6 - hoje.weekday()) % 7
    proximo_domingo = hoje + timedelta(days=dias_para_domingo)

    meses_ptbr = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                  'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    groups = OrderedDict()
    cur = proximo_domingo
    conn = get_db()
    while cur <= fim:
        date_str = cur.strftime("%Y-%m-%d")
        ata = conn.execute("SELECT * FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?", (date_str, session['user_id'])).fetchone()
        tema = ''
        discursantes = []
        ata_id = None
        # valores por discursante
        tema_1 = tema_2 = tema_3 = ''
        obs_1 = obs_2 = obs_3 = ''
        outros = ''
        if ata:
            ata_id = ata['id']
            sac = conn.execute("SELECT * FROM sacramental WHERE ata_id = ?", (ata_id,)).fetchone()
            if sac:
                tema = sac['tema'] or ''
                # Preferir colunas individuais quando existirem (compatibilidade com banco antigo)
                try:
                    if 'discursante_1' in sac.keys():
                        d1 = sac['discursante_1'] or ''
                        d2 = sac['discursante_2'] or ''
                        d3 = sac['ultimo_discursante'] or ''
                        discursantes = [d1, d2, d3]
                        tema_1 = sac.get('tema_1') or '' if isinstance(sac, dict) or True else ''
                        tema_2 = sac.get('tema_2') or '' if isinstance(sac, dict) or True else ''
                        tema_3 = sac.get('tema_ultimo') or '' if isinstance(sac, dict) or True else ''
                        obs_1 = sac.get('obs_1') or '' if isinstance(sac, dict) or True else ''
                        obs_2 = sac.get('obs_2') or '' if isinstance(sac, dict) or True else ''
                        obs_3 = sac.get('obs_ultimo') or '' if isinstance(sac, dict) or True else ''
                        outros = sac.get('outros') or '' if isinstance(sac, dict) or True else ''
                    else:
                        raw = json.loads(sac['discursantes']) if sac['discursantes'] else []
                        # Garantir exatamente 3 posições (preencher com strings vazias se faltar)
                        discursantes = [raw[i] if i < len(raw) else '' for i in range(3)]
                except Exception:
                    discursantes = ['', '', '']
        else:
            # Sem ata: mostrar 3 campos vazios
            discursantes = ['', '', '']

        # Inserir valores de tema/obs no item para popular o template
        # (esses valores podem estar vazios se não existirem no DB)
        item_tema_1 = tema_1
        item_tema_2 = tema_2
        item_tema_3 = tema_3
        item_obs_1 = obs_1
        item_obs_2 = obs_2
        item_obs_3 = obs_3
        item_outros = outros

        month_label = meses_ptbr[cur.month] + ' ' + str(cur.year)
        # Primeiro domingo do mês é reunião de testemunhos — não selecionamos discursantes
        is_testimony = (cur.day <= 7)
        item = {
            'date': date_str,
            'data_formatada': cur.strftime("%d/%m/%Y"),
            'tema': tema,
            'discursantes': discursantes,
            'ata_id': ata_id,
            'is_testimony': is_testimony,
            'tema_1': item_tema_1,
            'tema_2': item_tema_2,
            'tema_3': item_tema_3,
            'obs_1': item_obs_1,
            'obs_2': item_obs_2,
            'obs_3': item_obs_3,
            'outros': item_outros
        }
        if month_label not in groups:
            groups[month_label] = []
        groups[month_label].append(item)

        cur += timedelta(days=7)
    conn.close()

    grouped = [{'month_label': m, 'entries': items} for m, items in groups.items()]
    return render_template('discursantes_temas.html', groups=grouped)

# Polling-based UI (no websockets) — same data, different template
@app.route('/discursantes_temas/polling', methods=['GET'])
@login_required
def discursantes_temas_polling():
    hoje = datetime.now().date()

    months_to_add = 3
    target_month = (hoje.month - 1 + months_to_add) % 12 + 1
    target_year = hoje.year + (hoje.month - 1 + months_to_add) // 12
    last_day = calendar.monthrange(target_year, target_month)[1]
    fim = datetime(year=target_year, month=target_month, day=last_day).date()

    dias_para_domingo = (6 - hoje.weekday()) % 7
    proximo_domingo = hoje + timedelta(days=dias_para_domingo)

    meses_ptbr = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                  'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    groups = OrderedDict()
    cur = proximo_domingo
    conn = get_db()
    while cur <= fim:
        date_str = cur.strftime("%Y-%m-%d")
        ata = conn.execute("SELECT * FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?", (date_str, session['user_id'])).fetchone()
        tema = ''
        discursantes = []
        ata_id = None
        tema_1 = tema_2 = tema_3 = ''
        obs_1 = obs_2 = obs_3 = ''
        outros = ''
        if ata:
            ata_id = ata['id']
            sac = conn.execute('SELECT * FROM sacramental WHERE ata_id = ?', (ata_id,)).fetchone()
            if sac:
                # Converter para dict para usar .get() de forma segura
                try:
                    sac = dict(sac)
                except Exception:
                    pass
                tema = sac.get('tema') or ''
                try:
                    if 'discursante_1' in sac.keys():
                        d1 = sac.get('discursante_1') or ''
                        d2 = sac.get('discursante_2') or ''
                        d3 = sac.get('ultimo_discursante') or ''
                        discursantes = [d1, d2, d3]
                        tema_1 = sac.get('tema_1') or ''
                        tema_2 = sac.get('tema_2') or ''
                        tema_3 = sac.get('tema_ultimo') or ''
                        obs_1 = sac.get('obs_1') or ''
                        obs_2 = sac.get('obs_2') or ''
                        obs_3 = sac.get('obs_ultimo') or ''
                        outros = sac.get('outros') or ''
                    else:
                        raw = json.loads(sac.get('discursantes') or '[]') if sac.get('discursantes') else []
                        discursantes = [raw[i] if i < len(raw) else '' for i in range(3)]
                except Exception:
                    discursantes = ['', '', '']
        else:
            discursantes = ['', '', '']

        item = {
            'date': date_str,
            'data_formatada': cur.strftime("%d/%m/%Y"),
            'tema': tema,
            'discursantes': discursantes,
            'ata_id': ata_id,
            'is_testimony': (cur.day <= 7),
            'tema_1': tema_1,
            'tema_2': tema_2,
            'tema_3': tema_3,
            'obs_1': obs_1,
            'obs_2': obs_2,
            'obs_3': obs_3,
            'outros': outros
        }
        month_label = meses_ptbr[cur.month] + ' ' + str(cur.year)
        if month_label not in groups:
            groups[month_label] = []
        groups[month_label].append(item)
        cur += timedelta(days=7)

    conn.close()
    grouped = [{'month_label': m, 'entries': items} for m, items in groups.items()]
    return render_template('discursantes_temas_polling.html', groups=grouped)


@app.route('/discursantes_temas/salvar', methods=['POST'])
@login_required
def salvar_discursantes_temas():
    date = request.form.get('date')
    tema = (request.form.get('tema') or '').strip()
    # Recebe campos individuais
    d1 = (request.form.get('discursante_1') or '').strip()
    d2 = (request.form.get('discursante_2') or '').strip()
    d3 = (request.form.get('discursante_3') or '').strip()  # tratado como último
    outros = (request.form.get('outros') or '').strip()

    tema_1 = (request.form.get('tema_1') or '').strip()
    tema_2 = (request.form.get('tema_2') or '').strip()
    tema_3 = (request.form.get('tema_3') or '').strip()
    obs_1 = (request.form.get('obs_1') or '').strip()
    obs_2 = (request.form.get('obs_2') or '').strip()
    obs_3 = (request.form.get('obs_3') or '').strip()

    try:
        dt = datetime.strptime(date, "%Y-%m-%d").date()
    except Exception:
        flash('Data inválida', 'error')
        return redirect(url_for('discursantes_temas'))

    hoje = datetime.now().date()
    fim = hoje + timedelta(days=90)
    # Bloquear salvamento em primeiros domingos (reunião de testemunhos)
    if dt.day <= 7:
        flash('Primeiro domingo do mês é reunião de testemunhos; não é possível selecionar discursantes.', 'error')
        return redirect(url_for('discursantes_temas'))
    if dt < hoje or dt > fim or dt.weekday() != 6:
        flash('Data fora do intervalo permitido', 'error')
        return redirect(url_for('discursantes_temas'))

    conn = get_db()
    ata = conn.execute("SELECT * FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?", (date, session['user_id'])).fetchone()
    if not ata:
        cur = conn.execute("INSERT INTO atas (data, tipo, ala_id) VALUES (?, 'sacramental', ?)", (date, session['user_id']))
        ata_id = cur.lastrowid
    else:
        ata_id = ata['id']

    sac = conn.execute("SELECT * FROM sacramental WHERE ata_id = ?", (ata_id,)).fetchone()
    ultimo = d3  # Terceiro discursante será salvo como 'ultimo_discursante'
    if sac:
        conn.execute("UPDATE sacramental SET tema = ?, discursante_1 = ?, discursante_2 = ?, outros = ?, tema_1 = ?, tema_2 = ?, tema_ultimo = ?, obs_1 = ?, obs_2 = ?, obs_ultimo = ?, ultimo_discursante = ? WHERE ata_id = ?", (tema, d1, d2, outros, tema_1, tema_2, tema_3, obs_1, obs_2, obs_3, ultimo, ata_id))
    else:
        conn.execute("INSERT INTO sacramental (ata_id, tema, discursante_1, discursante_2, outros, tema_1, tema_2, tema_ultimo, obs_1, obs_2, obs_ultimo, ultimo_discursante) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (ata_id, tema, d1, d2, outros, tema_1, tema_2, tema_3, obs_1, obs_2, obs_3, ultimo))
    conn.commit()

    # Emitir atualizações via websocket para sincronizar outras abas/páginas
    try:
        room_date = f"date-{date}"
        room_ata = f"ata-{ata_id}" if ata_id else None
        updates = {
            'discursante_1': d1,
            'discursante_2': d2,
            'discursante_3': d3,
            'tema': tema,
            'tema_1': tema_1,
            'tema_2': tema_2,
            'tema_3': tema_3,
            'obs_1': obs_1,
            'obs_2': obs_2,
            'obs_3': obs_3
        }
        for name, val in updates.items():
            print(f"[socket][emit] (disc_temas) to {room_date} name={name} value={val!r} date={date}")
            # Se a emissão é originada por uma requisição HTTP padrão, o objeto `request` não terá `sid`.
            # Nesse caso usamos include_self=True para evitar que o Flask-SocketIO tente acessar request.sid.
            emit_include_self = True if not hasattr(request, 'sid') else False
            socketio.emit('field_update', {'ata_id': room_date, 'date': date, 'name': name, 'value': val}, to=room_date, include_self=emit_include_self)
            if room_ata:
                print(f"[socket][emit] (disc_temas) to {room_ata} name={name} value={val!r} date={date}")
                socketio.emit('field_update', {'ata_id': room_ata, 'date': date, 'name': name, 'value': val}, to=room_ata, include_self=emit_include_self)
    except Exception as e:
        print(f"[socket] erro ao emitir atualizacoes (discursantes_temas): {e}")
    conn.close()

    # If this is an AJAX request from the auto-save client, return JSON with the saved values
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        saved = {
            'discursante_1': d1,
            'discursante_2': d2,
            'discursante_3': d3,
            'tema': tema,
            'tema_1': tema_1,
            'tema_2': tema_2,
            'tema_3': tema_3,
            'obs_1': obs_1,
            'obs_2': obs_2,
            'obs_3': obs_3,
            'date': date,
            'ata_id': ata_id
        }
        return jsonify(saved)

    flash('Discursantes e tema salvos com sucesso', 'success')
    return redirect(url_for('discursantes_temas'))


# Rota para editar uma ata existente
@app.route("/ata/editar/<int:ata_id>")
@login_required
def editar_ata(ata_id):
    """Rota para editar uma ata existente"""
    conn = get_db()
    ata = conn.execute(
        "SELECT * FROM atas WHERE id=? AND ala_id=?", 
        (ata_id, session['user_id'])
    ).fetchone()
    
    if not ata:
        flash("Ata não encontrada ou você não tem permissão para editá-la.", "error")
        return redirect(url_for('index'))
    
    # Redireciona para o formulário apropriado com os dados existentes
    if ata["tipo"] == "sacramental":
        return redirect(url_for("form_ata", tipo="sacramental", data=ata["data"], editar=ata_id))
    else:
        return redirect(url_for("form_ata", tipo="batismo", data=ata["data"], editar=ata_id))

# Rota para excluir uma ata
@app.route("/ata/excluir/<int:ata_id>")
@login_required
def excluir_ata(ata_id: int):
    """Rota para excluir uma ata"""
    conn = get_db()
    
    # Primeiro, exclui os detalhes específicos
    ata = conn.execute("SELECT * FROM atas WHERE id=?", (ata_id,)).fetchone()
    if ata:
        if ata["tipo"] == "sacramental":
            conn.execute("DELETE FROM sacramental WHERE ata_id=?", (ata_id,))
        else:
            conn.execute("DELETE FROM batismo WHERE ata_id=?", (ata_id,))
        
        # Depois exclui a ata principal
        conn.execute("DELETE FROM atas WHERE id=?", (ata_id,))
        conn.commit()
        flash("Ata excluída com sucesso!", "success")
    else:
        flash("Ata não encontrada", "error")
    
    # Always return a redirect response
    return redirect(url_for("index"))

# Rota para listar atas por mês
@app.route("/atas/mes/<string:mes>")
@login_required
def listar_atas_mes(mes):
    conn = get_db()
    
    try:
        # Validar formato do mês (YYYY-MM)
        datetime.strptime(mes, "%Y-%m")
        
        atas = conn.execute(
            "SELECT * FROM atas WHERE strftime('%Y-%m', data) = ? AND ala_id = ? ORDER BY data DESC", 
            (mes, session['user_id'])
        ).fetchall()
        
        # Formatar nome do mês para exibição EM PORTUGUÊS
        meses_ptbr = [
            '', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        data_mes = datetime.strptime(mes, "%Y-%m")
        mes_nome = meses_ptbr[data_mes.month] + " " + str(data_mes.year)
        
        return render_template("_atas_list.html", 
                             atas=atas, 
                             mes_selecionado_nome=mes_nome)
    
    except ValueError:
        return "<div class='info-card'>Mês inválido.</div>"

# Filtro de template para carregar listas JSON
@app.template_filter('loads')
def json_loads_filter(s: str) -> list:
    """Template filter to parse JSON strings - always returns a list"""
    if not s:
        return []
    try:
        result = json.loads(s)
        # Ensure we always return a list, even if JSON contains other types
        if isinstance(result, list):
            return result
        else:
            return [result] if result is not None else []
    except (json.JSONDecodeError, TypeError, ValueError):
        return []

# Rota para criar nova ata
@app.route("/ata/nova", methods=["GET", "POST"])
@login_required
def nova_ata():
    if request.method == "POST":
        tipo = request.form.get("tipo")
        data = request.form.get("data")
        
        # Validação básica
        if not tipo or not data:
            flash("Erro: Tipo e data são obrigatórios", "error")
            return render_template("nova_ata.html")
            
        # Validação de data - APENAS VERIFICA SE É UMA DATA VÁLIDA
        try:
            datetime.strptime(data, "%Y-%m-%d")
        except ValueError:
            flash("Erro: Data inválida", "error")
            return render_template("nova_ata.html")
            
        # Se for sacramental e já existir uma ata para a mesma data, abrir para edição
        if tipo == 'sacramental':
            conn = get_db()
            existing = conn.execute("SELECT id FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?", (data, session['user_id'])).fetchone()
            conn.close()
            if existing:
                flash('Já existe uma ata sacramental para essa data. Abrindo para edição.', 'info')
                return redirect(url_for('editar_ata', ata_id=existing['id']))

        return redirect(url_for("form_ata", tipo=tipo, data=data))
    
    # Data padrão: próximo domingo ou hoje se for domingo
    hoje = datetime.now().date()
    dias_para_domingo = (6 - hoje.weekday()) % 7
    if dias_para_domingo == 0:  # Se hoje é domingo
        data_padrao = hoje.strftime("%Y-%m-%d")
    else:
        data_padrao = (hoje + timedelta(days=dias_para_domingo)).strftime("%Y-%m-%d")
    
    return render_template("nova_ata.html", data_padrao=data_padrao)

# Rota para formulário de ata (criação/edição)
@app.route("/ata/form", methods=["GET", "POST"])
@login_required
def form_ata():
    if request.method == "POST":
        tipo = request.form.get("tipo")
        data = request.form.get("data")
        ata_id_editar = request.form.get("editar")
        
        # Validação básica
        if not tipo or not data:
            flash("Erro: Tipo e data são obrigatórios", "error")
            return redirect(url_for('nova_ata'))
        
        # Validação de data
        try:
            datetime.strptime(data, "%Y-%m-%d")
        except ValueError:
            flash("Erro: Data inválida", "error")
            return redirect(url_for('nova_ata'))
        
        conn = get_db()
        
        if ata_id_editar:
            # Modo edição - verificar se a ata pertence à ala do usuário
            ata_existente = conn.execute(
                "SELECT * FROM atas WHERE id = ? AND ala_id = ?", 
                (ata_id_editar, session['user_id'])
            ).fetchone()
            
            if not ata_existente:
                flash("Você não tem permissão para editar esta ata.", "error")
                return redirect(url_for('index'))
            
            # Atualiza a ata existente
            conn.execute("UPDATE atas SET tipo=?, data=? WHERE id=?", (tipo, data, ata_id_editar))
            ata_id = ata_id_editar
        else:
            # Modo criação - insere nova ata com ala_id
            conn.execute(
                "INSERT INTO atas (tipo, data, ala_id) VALUES (?, ?, ?)", 
                (tipo, data, session['user_id'])
            )
            ata_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        if tipo == "sacramental":
            raw_discursantes = request.form.getlist("discursantes[]")
            # Preserve order and ensure exactly 2 positions (pad with empty strings)
            discursantes = [ (raw_discursantes[i].strip() if i < len(raw_discursantes) and raw_discursantes[i] and raw_discursantes[i].strip() else '') for i in range(2) ]
            
            raw_anuncios = request.form.getlist("anuncios[]")
            # Support multiline textarea: split entries by newline and trim
            anuncios = []
            for a in raw_anuncios:
                if a and a.strip():
                    for l in a.splitlines():
                        if l and l.strip():
                            anuncios.append(l.strip())
            
            # Convert multi-name fields from arrays into JSON strings (maintain empty string for compatibility)
            def _collect(name):
                arr = [v.strip() for v in request.form.getlist(f'{name}[]') if v and v.strip()]
                if not arr:
                    raw = request.form.get(name)
                    if raw:
                        arr = [l.strip() for l in raw.splitlines() if l and l.strip()]
                return arr

            desobrigacoes_list = _collect('desobrigacoes')
            apoios_list = _collect('apoios')
            confirmacoes_list = _collect('confirmacoes_batismo')
            apoio_membros_list = _collect('apoio_membros')
            bencao_list = _collect('bencao_criancas')
            reconhecemos_list = _collect('reconhecemos_presenca')

            detalhes = {
                "presidido": request.form.get("presidido", ""),
                "dirigido": request.form.get("dirigido", ""),
                "recepcionistas": request.form.get("recepcionista", ""),
                "tema": request.form.get("tema", ""), 
                "pianista": request.form.get("pianista", ""),
                "regente_musica": request.form.get("regente_musica", ""),
                "reconhecemos_presenca": json.dumps(reconhecemos_list) if reconhecemos_list else '',
                "anuncios": anuncios,
                "hino_abertura": request.form.get("hino_abertura", ""),
                "oracao_abertura": request.form.get("oracao_abertura", ""),
                "desobrigacoes": json.dumps(desobrigacoes_list) if desobrigacoes_list else '',
                "apoios": json.dumps(apoios_list) if apoios_list else '',
                "confirmacoes_batismo": json.dumps(confirmacoes_list) if confirmacoes_list else '',
                "apoio_membros": json.dumps(apoio_membros_list) if apoio_membros_list else '',
                "bencao_criancas": json.dumps(bencao_list) if bencao_list else '',
                "hino_sacramental": request.form.get("hino_sacramental", ""),                "hino_intermediario": request.form.get("hino_intermediario", ""),
                "ultimo_discursante": request.form.get("ultimo_discursante", ""),  # NOVO
                "hino_encerramento": request.form.get("hino_encerramento", ""),
                "oracao_encerramento": request.form.get("oracao_encerramento", ""),
                "discursante_1": request.form.get("discursante_1", ""),
                "discursante_2": request.form.get("discursante_2", ""),
                "outros": request.form.get("outros", ""),
                "tema_1": request.form.get("tema_1", ""),
                "tema_2": request.form.get("tema_2", ""),
                "tema_ultimo": request.form.get("tema_ultimo", ""),
                "obs_1": request.form.get("obs_1", ""),
                "obs_2": request.form.get("obs_2", ""),
                "obs_ultimo": request.form.get("obs_ultimo", "")
            }
            
            try:
                if ata_id_editar:
                    # Atualiza registro existente COM TEMA e colunas individuais
                    # Proteção: não sobrescrever valores existentes com campos vazios enviados pelo formulário
                    existing_row = conn.execute("SELECT * FROM sacramental WHERE ata_id = ?", (ata_id,)).fetchone()
                    if existing_row:
                        existing = dict(existing_row)
                        for _key in [
                            'presidido','dirigido','recepcionistas','pianista','regente_musica',
                            'reconhecemos_presenca','anuncios','hinos','oracoes',
                            'discursante_1','discursante_2','outros','tema_1','tema_2','tema_ultimo','obs_1','obs_2','obs_ultimo',
                            'hino_sacramental','hino_intermediario','desobrigacoes','apoios',
                            'confirmacoes_batismo','apoio_membros','bencao_criancas','ultimo_discursante','tema'
                        ]:
                            try:
                                v = detalhes.get(_key)
                            except Exception:
                                v = None
                            # Se o formulário não enviou valor (string vazia ou None ou empty list), manter existente
                            if (v is None or v == '' or (isinstance(v, list) and len(v) == 0)) and existing.get(_key):
                                detalhes[_key] = existing.get(_key)

                    conn.execute("""
                        UPDATE sacramental 
                        SET presidido=?, dirigido=?, recepcionistas=?, pianista=?, regente_musica=?, 
                            reconhecemos_presenca=?, anuncios=?, hinos=?, oracoes=?, 
                            discursante_1=?, discursante_2=?, outros=?, tema_1=?, tema_2=?, tema_ultimo=?, obs_1=?, obs_2=?, obs_ultimo=?,
                            hino_sacramental=?, hino_intermediario=?, desobrigacoes=?, apoios=?, 
                            confirmacoes_batismo=?, apoio_membros=?, bencao_criancas=?, ultimo_discursante=?, tema=?
                        WHERE ata_id=?
                    """, (
                        detalhes["presidido"], 
                        detalhes["dirigido"],
                        detalhes["recepcionistas"],
                        detalhes["pianista"],
                        detalhes["regente_musica"],
                        detalhes["reconhecemos_presenca"],
                        json.dumps(detalhes["anuncios"]),
                        json.dumps([detalhes["hino_abertura"], detalhes["hino_encerramento"]]), 
                        json.dumps([detalhes["oracao_abertura"], detalhes["oracao_encerramento"]]), 
                        detalhes["discursante_1"], detalhes["discursante_2"], detalhes.get("outros", ""), detalhes.get("tema_1", ""), detalhes.get("tema_2", ""), detalhes.get("tema_ultimo", ""), detalhes.get("obs_1", ""), detalhes.get("obs_2", ""), detalhes.get("obs_ultimo", ""),
                        detalhes["hino_sacramental"],
                        detalhes["hino_intermediario"],
                        detalhes["desobrigacoes"],
                        detalhes["apoios"],
                        detalhes["confirmacoes_batismo"],
                        detalhes["apoio_membros"],
                        detalhes["bencao_criancas"],
                        detalhes["ultimo_discursante"],
                        detalhes["tema"],  # ← ADICIONAR AQUI
                        ata_id
                    ))
                else:
                    # Insere novo registro COM TEMA e colunas individuais
                    conn.execute("""
                        INSERT INTO sacramental (ata_id, presidido, dirigido, recepcionistas, pianista, regente_musica, 
                            reconhecemos_presenca, anuncios, hinos, oracoes, discursante_1, discursante_2, outros, tema_1, tema_2, tema_ultimo, obs_1, obs_2, obs_ultimo, hino_sacramental, hino_intermediario,
                            desobrigacoes, apoios, confirmacoes_batismo, apoio_membros, bencao_criancas, ultimo_discursante, tema) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        ata_id, 
                        detalhes["presidido"], 
                        detalhes["dirigido"],
                        detalhes["recepcionistas"],
                        detalhes["pianista"],
                        detalhes["regente_musica"],
                        detalhes["reconhecemos_presenca"],
                        json.dumps(detalhes["anuncios"]),
                        json.dumps([detalhes["hino_abertura"], detalhes["hino_encerramento"]]), 
                        json.dumps([detalhes["oracao_abertura"], detalhes["oracao_encerramento"]]), 
                        detalhes.get("discursante_1", ""), detalhes.get("discursante_2", ""), detalhes.get("outros", ""), detalhes.get("tema_1", ""), detalhes.get("tema_2", ""), detalhes.get("tema_ultimo", ""), detalhes.get("obs_1", ""), detalhes.get("obs_2", ""), detalhes.get("obs_ultimo", ""),
                        detalhes["hino_sacramental"],
                        detalhes["hino_intermediario"],
                        detalhes["desobrigacoes"],
                        detalhes["apoios"],
                        detalhes["confirmacoes_batismo"],
                        detalhes["apoio_membros"],
                        detalhes["bencao_criancas"],
                        detalhes["ultimo_discursante"],
                        detalhes["tema"]  # ← ADICIONAR AQUI
                    ))
            except sqlite3.OperationalError as e:
                conn.rollback()
                conn.close()
                flash(f"Erro ao salvar ata (DB): {e}", "error")
                return redirect(url_for('form_ata', tipo=tipo, data=data))
            except Exception as e:
                conn.rollback()
                conn.close()
                flash(f"Erro ao salvar ata: {e}", "error")
                return redirect(url_for('form_ata', tipo=tipo, data=data))
        
        elif tipo == "batismo":
            batizados = request.form.getlist("batizados[]")
            # Filtrar batizados vazios
            batizados = [b for b in batizados if b and b.strip()]
            
            detalhes = {
                "presidido": request.form.get("presidido", ""),
                "dirigido": request.form.get("dirigido", ""),
                "dedicado": request.form.get("dedicado", ""),
                "testemunha1": request.form.get("testemunha1", ""),
                "testemunha2": request.form.get("testemunha2", ""),
                "batizados": batizados
            }
            
            if ata_id_editar:
                # Atualiza registro existente
                conn.execute("""
                    UPDATE batismo 
                    SET dedicado=?, presidido=?, dirigido=?, batizados=?, testemunha1=?, testemunha2=? 
                    WHERE ata_id=?
                """, (
                    detalhes["dedicado"], 
                    detalhes["presidido"], 
                    detalhes["dirigido"], 
                    json.dumps(detalhes["batizados"]), 
                    detalhes["testemunha1"], 
                    detalhes["testemunha2"], 
                    ata_id
                ))
            else:
                # Insere novo registro
                conn.execute("""
                    INSERT INTO batismo (ata_id, dedicado, presidido, dirigido, batizados, testemunha1, testemunha2) 
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    ata_id, 
                    detalhes["dedicado"], 
                    detalhes["presidido"], 
                    detalhes["dirigido"], 
                    json.dumps(detalhes["batizados"]), 
                    detalhes["testemunha1"], 
                    detalhes["testemunha2"]
                ))
        
        conn.commit()

        # Notificar clientes conectados (Discursantes & Temas) para sincronizar campos
        try:
            room_ata = f"ata-{ata_id}"
            room_date = f"date-{data}"
            updates = {
                'discursante_1': detalhes.get('discursante_1', ''),
                'discursante_2': detalhes.get('discursante_2', ''),
                # O formulário de 'discursantes_temas' espera 'discursante_3' — mapeamos do 'ultimo_discursante'
                'discursante_3': detalhes.get('ultimo_discursante', ''),
                'tema': detalhes.get('tema', ''),
                'tema_1': detalhes.get('tema_1', ''),
                'tema_2': detalhes.get('tema_2', ''),
                'tema_3': detalhes.get('tema_ultimo', ''),
                'obs_1': detalhes.get('obs_1', ''),
                'obs_2': detalhes.get('obs_2', ''),
                'obs_3': detalhes.get('obs_ultimo', '')
            }
            # Determinar se devemos usar include_self=False — em requisições HTTP não existe request.sid
            emit_include_self = True if not hasattr(request, 'sid') else False
            for name, val in updates.items():
                # Emitir para sala por ATA e por DATA para cobrir ambos os casos
                print(f"[socket][emit] (form_ata) to={room_ata} / {room_date} name={name} value={val!r} date={data}")
                socketio.emit('field_update', {'ata_id': room_ata, 'date': data, 'name': name, 'value': val}, to=room_ata, include_self=emit_include_self)
                socketio.emit('field_update', {'ata_id': room_date, 'date': data, 'name': name, 'value': val}, to=room_date, include_self=emit_include_self)
        except Exception as e:
            print(f"[socket] erro ao emitir atualizacoes: {e}")
        conn.close()
        flash("Ata salva com sucesso!", "success")
        return redirect(url_for("visualizar_ata", ata_id=ata_id))

    # GET request
    tipo = request.args.get("tipo")
    data = request.args.get("data")
    editar = request.args.get("editar")
    
    # Se é sacramental e não está em modo editar, redirecionar para edição caso já exista ata
    if tipo == 'sacramental' and not editar:
        conn = get_db()
        existing = conn.execute("SELECT id FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?", (data, session['user_id'])).fetchone()
        conn.close()
        if existing:
            flash('Já existe uma ata para essa data. Abrindo para edição.', 'info')
            return redirect(url_for('editar_ata', ata_id=existing['id']))
    
    # Lógica para carregar dados existentes se estiver editando
    dados_existentes = {}
    if editar:
        conn = get_db()
        if tipo == "sacramental":
            dados = conn.execute("SELECT * FROM sacramental WHERE ata_id=?", (editar,)).fetchone()
            if dados:
                dados_existentes = dict(dados)
                # Converter JSON strings de volta para objetos
                if dados_existentes.get('hinos'):
                    hinos = json.loads(dados_existentes['hinos'])
                    dados_existentes['hino_abertura'] = hinos[0] if len(hinos) > 0 else ''
                    dados_existentes['hino_encerramento'] = hinos[1] if len(hinos) > 1 else ''
                if dados_existentes.get('oracoes'):
                    oracoes = json.loads(dados_existentes['oracoes'])
                    dados_existentes['oracao_abertura'] = oracoes[0] if len(oracoes) > 0 else ''
                    dados_existentes['oracao_encerramento'] = oracoes[1] if len(oracoes) > 1 else ''
                # Preencher nomes a partir do novo formato (colunas individuais) ou do formato antigo (JSON)
                if 'discursante_1' in dados_existentes:
                    dados_existentes['discursante_1'] = dados_existentes.get('discursante_1') or ''
                    dados_existentes['discursante_2'] = dados_existentes.get('discursante_2') or ''
                    dados_existentes['outros'] = dados_existentes.get('outros') or ''
                    dados_existentes['tema_1'] = dados_existentes.get('tema_1') or ''
                    dados_existentes['tema_2'] = dados_existentes.get('tema_2') or ''
                    dados_existentes['tema_ultimo'] = dados_existentes.get('tema_ultimo') or ''
                    dados_existentes['obs_1'] = dados_existentes.get('obs_1') or ''
                    dados_existentes['obs_2'] = dados_existentes.get('obs_2') or ''
                    dados_existentes['obs_ultimo'] = dados_existentes.get('obs_ultimo') or ''
                elif dados_existentes.get('discursantes'):
                    try:
                        parsed = json.loads(dados_existentes['discursantes'])
                    except:
                        parsed = []
                    dados_existentes['discursante_1'] = parsed[0] if len(parsed) > 0 else ''
                    dados_existentes['discursante_2'] = parsed[1] if len(parsed) > 1 else ''
                    dados_existentes['outros'] = ''
                    dados_existentes['tema_1'] = ''
                    dados_existentes['tema_2'] = ''
                    dados_existentes['tema_ultimo'] = ''
                    dados_existentes['obs_1'] = ''
                    dados_existentes['obs_2'] = ''

                # Compatibilidade: alias singular para campos usados no template (ex: recepcionista)
                # Alguns locais do template usam 'recepcionista' (singular) enquanto o DB usa 'recepcionistas' (plural)
                if 'recepcionistas' in dados_existentes and not dados_existentes.get('recepcionista'):
                    dados_existentes['recepcionista'] = dados_existentes.get('recepcionistas') or ''

                # Garantir chaves básicas sempre presentes para evitar None no template
                for _k in ['presidido','dirigido','recepcionistas','recepcionista','pianista','regente_musica','tema']:
                    dados_existentes[_k] = dados_existentes.get(_k) or ''
                    dados_existentes['obs_ultimo'] = ''
                # Converter anúncios (JSON -> lista)
                if dados_existentes.get('anuncios'):
                    try:
                        dados_existentes['anuncios'] = json.loads(dados_existentes['anuncios'])
                    except Exception:
                        dados_existentes['anuncios'] = str(dados_existentes['anuncios']).splitlines()

                # Converter reconhecemos_presenca (JSON -> lista ou splitlines)
                if dados_existentes.get('reconhecemos_presenca'):
                    try:
                        parsed = json.loads(dados_existentes['reconhecemos_presenca'])
                        if isinstance(parsed, list):
                            dados_existentes['reconhecemos_presenca'] = parsed
                        else:
                            dados_existentes['reconhecemos_presenca'] = str(dados_existentes['reconhecemos_presenca']).splitlines()
                    except Exception:
                        dados_existentes['reconhecemos_presenca'] = str(dados_existentes['reconhecemos_presenca']).splitlines()
                else:
                    dados_existentes['reconhecemos_presenca'] = []

                # Converter campos de AÇÕES (podem ser JSON lists ou textos antigos)
                for _fld in ['desobrigacoes','apoios','confirmacoes_batismo','apoio_membros','bencao_criancas']:
                    if dados_existentes.get(_fld):
                        try:
                            parsed = json.loads(dados_existentes[_fld])
                            if isinstance(parsed, list):
                                dados_existentes[_fld] = parsed
                            else:
                                dados_existentes[_fld] = str(dados_existentes[_fld]).splitlines()
                        except Exception:
                            dados_existentes[_fld] = str(dados_existentes[_fld]).splitlines()
                    else:
                        dados_existentes[_fld] = []
        else:
            dados = conn.execute("SELECT * FROM batismo WHERE ata_id=?", (editar,)).fetchone()
            if dados:
                dados_existentes = dict(dados)
                if dados_existentes.get('batizados'):
                    dados_existentes['batizados'] = json.loads(dados_existentes['batizados'])
    
    if not tipo or not data:
        flash("Erro: Tipo e data são obrigatórios", "error")
        return redirect(url_for("nova_ata"))
    
    if tipo == "sacramental":
        dt = datetime.strptime(data, "%Y-%m-%d")
        primeiro_domingo = min([d for d in range(1, 8) if calendar.weekday(dt.year, dt.month, d) == 6])
        is_primeiro_domingo = dt.day == primeiro_domingo
        

        discursantes_recentes = get_discursantes_recentes() 
        temas_recentes = get_temas_recentes() 
        hinos_recentes = get_hinos_recentes() 
        
        conn = get_db()
        unidade_row = conn.execute("SELECT * FROM unidades WHERE ala_id = ?", (session['user_id'],)).fetchone()
        estaca_row = None

        if unidade_row and unidade_row['estaca_id']:
            estaca_row = conn.execute("SELECT * FROM estacas WHERE id = ?", (unidade_row['estaca_id'],)).fetchone()

        unidade = dict(unidade_row) if unidade_row else {}
        estaca = dict(estaca_row) if estaca_row else {}

        return render_template("sacramental.html", 
                             primeiro=is_primeiro_domingo, 
                             data=data, 
                             editar=editar, 
                             dados=dados_existentes,
                             discursantes_recentes=discursantes_recentes,
                             temas_recentes=temas_recentes,
                             hinos_recentes=hinos_recentes,
                             unidade=unidade,
                             estaca=estaca)
    elif tipo == "batismo":
        return render_template("batismo.html", 
                             data=data, 
                             editar=editar, 
                             dados=dados_existentes)
    else:
        flash("Tipo de ata não reconhecido", "error")
        return redirect(url_for("nova_ata"))

# Rota para visualizar uma ata selecionada
@app.route("/ata/<int:ata_id>")
@login_required
def visualizar_ata(ata_id):
    conn = get_db()
    ata = conn.execute(
        "SELECT * FROM atas WHERE id=? AND ala_id=?", 
        (ata_id, session['user_id'])
    ).fetchone()
    
    if not ata:
        flash("Ata não encontrada ou você não tem permissão para visualizá-la.", "error")
        return redirect(url_for("index"))
        
    # Buscar template padrão para sacramental
    template = None
    if ata["tipo"] == "sacramental":
        # Tente diferentes formas de buscar o template
        template = conn.execute(
            "SELECT * FROM templates WHERE nome = 'Sacramental Padrão'"
        ).fetchone()
        
        if not template:
            template = conn.execute(
                "SELECT * FROM templates WHERE tipo_template = 1"
            ).fetchone()
        
        if template:
            template = dict(template)
            print(f"DEBUG: Template carregado - {template.get('nome', 'Sem nome')}")
    
    if ata["tipo"] == "sacramental":
        detalhes = conn.execute("SELECT * FROM sacramental WHERE ata_id=?", (ata_id,)).fetchone()
        if detalhes:
            # Converter para dicionário para facilitar o acesso
            detalhes_dict = dict(detalhes)
            if detalhes_dict.get('hinos'):
                try:
                    hinos = json.loads(detalhes_dict['hinos'])
                    detalhes_dict['hino_abertura'] = hinos[0] if len(hinos) > 0 else ''
                    detalhes_dict['hino_encerramento'] = hinos[1] if len(hinos) > 1 else ''
                except:
                    detalhes_dict['hino_abertura'] = ''
                    detalhes_dict['hino_encerramento'] = ''
                    
            if detalhes_dict.get('oracoes'):
                try:
                    oracoes = json.loads(detalhes_dict['oracoes'])
                    detalhes_dict['oracao_abertura'] = oracoes[0] if len(oracoes) > 0 else ''
                    detalhes_dict['oracao_encerramento'] = oracoes[1] if len(oracoes) > 1 else ''
                except:
                    detalhes_dict['oracao_abertura'] = ''
                    detalhes_dict['oracao_encerramento'] = ''
                    
            if detalhes_dict.get('discursantes'):
                try:
                    detalhes_dict['discursantes'] = json.loads(detalhes_dict['discursantes'])
                except:
                    detalhes_dict['discursantes'] = []
                    
            if detalhes_dict.get('anuncios'):
                try:
                    detalhes_dict['anuncios'] = json.loads(detalhes_dict['anuncios'])
                except:
                    detalhes_dict['anuncios'] = []

            # Convertemos reconhecemos_presenca se estiver em JSON
            if detalhes_dict.get('reconhecemos_presenca'):
                try:
                    parsed = json.loads(detalhes_dict['reconhecemos_presenca'])
                    if isinstance(parsed, list):
                        detalhes_dict['reconhecemos_presenca'] = parsed
                    else:
                        detalhes_dict['reconhecemos_presenca'] = str(detalhes_dict['reconhecemos_presenca']).splitlines()
                except Exception:
                    detalhes_dict['reconhecemos_presenca'] = str(detalhes_dict['reconhecemos_presenca']).splitlines()
            else:
                detalhes_dict['reconhecemos_presenca'] = []
                    
            detalhes = detalhes_dict
        else:
            detalhes = {}
    else:
        detalhes = conn.execute("SELECT * FROM batismo WHERE ata_id=?", (ata_id,)).fetchone()
        if detalhes:
            detalhes_dict = dict(detalhes)
            if detalhes_dict.get('batizados'):
                try:
                    detalhes_dict['batizados'] = json.loads(detalhes_dict['batizados'])
                except:
                    detalhes_dict['batizados'] = []
            detalhes = detalhes_dict
        else:
            detalhes = {}
    
    # Ler os textos padrão dos convites (1º, 2º e 3º discursante)
    def _read_discursante_text(n):
        try:
            path_txt = os.path.join(app.root_path, "templates", "texts", f"discursante_{n}.txt")
            with open(path_txt, "r", encoding="utf-8") as f:
                return f.read()
        except Exception as e:
            print(f"Aviso: não foi possível ler discursante_{n}.txt ->", e)
            return ""

    discursante_1_text = _read_discursante_text(1)
    discursante_2_text = _read_discursante_text(2)
    discursante_3_text = _read_discursante_text(3)

    # Se a estrutura antiga já foi atualizada para colunas individuais, garantir campos derivados
    if isinstance(detalhes, dict) and 'discursante_1' in detalhes:
        detalhes['discursantes'] = [detalhes.get('discursante_1') or '', detalhes.get('discursante_2') or '', detalhes.get('ultimo_discursante') or '']
        detalhes['tema_1'] = detalhes.get('tema_1') or ''
        detalhes['tema_2'] = detalhes.get('tema_2') or ''
        detalhes['tema_ultimo'] = detalhes.get('tema_ultimo') or ''
        detalhes['obs_1'] = detalhes.get('obs_1') or ''
        detalhes['obs_2'] = detalhes.get('obs_2') or ''
        detalhes['obs_ultimo'] = detalhes.get('obs_ultimo') or ''

        # Parse JSON lists for action fields (backwards-compatible with plain strings)
        for _fld in ['desobrigacoes','apoios','confirmacoes_batismo','apoio_membros','bencao_criancas']:
            if detalhes.get(_fld):
                try:
                    detalhes[_fld] = json.loads(detalhes[_fld])
                except Exception:
                    # keep as string, template will splitlines when needed
                    pass

    conn.close()

    return render_template(
        "visualizar_ata.html",
        ata=ata,
        detalhes=detalhes,
        template=template,
        discursante_1_text=discursante_1_text,
        discursante_2_text=discursante_2_text,
        discursante_3_text=discursante_3_text
    )

# Rota para exportar ata como PDF simples
@app.route("/ata/exportar/<int:ata_id>")
@login_required
def exportar_pdf(ata_id):
    from functions.pdf_exporters import exportar_pdf_bytes
    
    conn = get_db() 
    try:
        # Usamos um cursor explícito para maior robustez
        cursor = conn.cursor() 
        
        # 1. Buscar a Ata
        cursor.execute(
            "SELECT * FROM atas WHERE id=? AND ala_id=?", 
            (ata_id, session['user_id'])
        )
        ata = cursor.fetchone()
        
        if not ata:
            raise ValueError("Ata não encontrada")
        
        ata = dict(ata)

        # Incluir o nome da ala (unidade) para que o PDF possa exibir 'Ala Nome' no cabeçalho
        cursor.execute("SELECT nome FROM unidades WHERE ala_id = ? LIMIT 1", (ata.get('ala_id'),))
        unidade = cursor.fetchone()
        if unidade:
            ata['ala_nome'] = unidade['nome']
        
        # =========================================================================
        # CORREÇÃO: Buscar o Template Padrão (ID 1), pois a tabela templates 
        # não possui a coluna ala_id.
        # =========================================================================
        cursor.execute(
            # Anteriormente: "SELECT * FROM templates WHERE ala_id=? LIMIT 1"
            "SELECT * FROM templates WHERE id=1 LIMIT 1" # Agora busca o template padrão (ID 1)
        )
        template = cursor.fetchone()
        
        if template:
            template = dict(template)
        else:
            template = {}
        
        
        # 4. Buscar detalhes conforme tipo (Lógica de deserialização mantida)
        if ata["tipo"] == "sacramental":
            cursor.execute("SELECT * FROM sacramental WHERE ata_id=?", (ata_id,))
            detalhes = cursor.fetchone()
            
            if detalhes:
                detalhes_dict = dict(detalhes)
                
                # Deserialização de JSON
                keys_to_load = ['hinos', 'oracoes', 'discursantes', 'anuncios', 
                                'desobrigacoes', 'apoios', 'confirmacoes_batismo', 
                                'apoio_membro_novo', 'bencao_crianca']

                for key in keys_to_load:
                    if detalhes_dict.get(key) and isinstance(detalhes_dict[key], str):
                        try:
                            detalhes_dict[key] = json.loads(detalhes_dict[key])
                        except:
                            if key in ['discursantes', 'anuncios', 'desobrigacoes', 'apoios', 'confirmacoes_batismo', 'apoio_membro_novo', 'bencao_crianca']:
                                detalhes_dict[key] = []
                            pass

                # Construir lista de discursantes a partir das colunas individuais quando presentes
                if 'discursante_1' in detalhes_dict:
                    d1 = detalhes_dict.get('discursante_1') or ''
                    d2 = detalhes_dict.get('discursante_2') or ''
                    d3 = detalhes_dict.get('ultimo_discursante') or ''
                    detalhes_dict['discursantes'] = [d1, d2, d3]
                    detalhes_dict['tema_1'] = detalhes_dict.get('tema_1') or ''
                    detalhes_dict['tema_2'] = detalhes_dict.get('tema_2') or ''
                    detalhes_dict['tema_ultimo'] = detalhes_dict.get('tema_ultimo') or ''
                    detalhes_dict['obs_1'] = detalhes_dict.get('obs_1') or ''
                    detalhes_dict['obs_2'] = detalhes_dict.get('obs_2') or ''
                    detalhes_dict['obs_ultimo'] = detalhes_dict.get('obs_ultimo') or ''

                # Tratamento específico de hinos e orações
                if isinstance(detalhes_dict.get('hinos'), list):
                    hinos = detalhes_dict['hinos']
                    detalhes_dict['hino_abertura'] = hinos[0] if len(hinos) > 0 else ''
                    detalhes_dict['hino_encerramento'] = hinos[1] if len(hinos) > 1 else ''
                    
                if isinstance(detalhes_dict.get('oracoes'), list):
                    oracoes = detalhes_dict['oracoes']
                    detalhes_dict['oracao_abertura'] = oracoes[0] if len(oracoes) > 0 else ''
                    detalhes_dict['oracao_encerramento'] = oracoes[1] if len(oracoes) > 1 else ''
                    
                detalhes = detalhes_dict
            else:
                detalhes = {}
        else: # Tipo batismo
            cursor.execute("SELECT * FROM batismo WHERE ata_id=?", (ata_id,))
            detalhes = cursor.fetchone()
            if detalhes:
                detalhes_dict = dict(detalhes)
                if detalhes_dict.get('batizados'):
                    try:
                        detalhes_dict['batizados'] = json.loads(detalhes_dict['batizados'])
                    except:
                        detalhes_dict['batizados'] = []
                detalhes = detalhes_dict
            else:
                detalhes = {}
        
        # 5. Converter para PDF
        buffer, filename, mimetype = exportar_pdf_bytes(ata, detalhes, template, filename=f"ata_{ata_id}.pdf")
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype=mimetype)
        
    except Exception as e:
        print(f"======== ERRO CRÍTICO NA EXPORTAÇÃO DE PDF: {e} ========")
        flash(f"Erro ao exportar PDF: {str(e)}", "error")
        return redirect(url_for("visualizar_ata", ata_id=ata_id))
    finally:
        conn.close()

# Rota para exportar ata como PDF SIMPLES (SOMENTE CAMPOS/SEM TEXTOS)
@app.route("/ata/exportar_simples/<int:ata_id>")
@login_required
def exportar_pdf_simples(ata_id):
    from functions.pdf_exporters import exportar_pdf_bytes
    
    conn = get_db() 
    try:
        cursor = conn.cursor() 
        
        # 1. Buscar a Ata
        cursor.execute(
            "SELECT * FROM atas WHERE id=? AND ala_id=?", 
            (ata_id, session['user_id'])
        )
        ata = cursor.fetchone()
        
        if not ata:
            raise ValueError("Ata não encontrada")
        
        ata = dict(ata)

        # Incluir o nome da ala (unidade) para exibição no cabeçalho do PDF
        cursor.execute("SELECT nome FROM unidades WHERE ala_id = ? LIMIT 1", (ata.get('ala_id'),))
        unidade = cursor.fetchone()
        if unidade:
            ata['ala_nome'] = unidade['nome']
        
        # 2. NÃO BUSCAR O TEMPLATE: template = {} ou template = None
        template = {} 
        
        # 3. Buscar detalhes conforme tipo (Lógica de deserialização mantida)
        if ata["tipo"] == "sacramental":
            cursor.execute("SELECT * FROM sacramental WHERE ata_id=?", (ata_id,))
            detalhes = cursor.fetchone()
            
            if detalhes:
                detalhes_dict = dict(detalhes)
                
                # Deserialização de JSON (simplificada)
                keys_to_load = ['hinos', 'oracoes', 'discursantes', 'anuncios', 
                                'desobrigacoes', 'apoios', 'confirmacoes_batismo', 
                                'apoio_membro_novo', 'bencao_crianca']

                for key in keys_to_load:
                    if detalhes_dict.get(key) and isinstance(detalhes_dict[key], str):
                        try:
                            detalhes_dict[key] = json.loads(detalhes_dict[key])
                        except:
                            if key in ['discursantes', 'anuncios', 'desobrigacoes', 'apoios', 'confirmacoes_batismo', 'apoio_membro_novo', 'bencao_crianca']:
                                detalhes_dict[key] = []
                            pass

                # Tratamento específico de hinos e orações
                if isinstance(detalhes_dict.get('hinos'), list):
                    hinos = detalhes_dict['hinos']
                    detalhes_dict['hino_abertura'] = hinos[0] if len(hinos) > 0 else ''
                    detalhes_dict['hino_encerramento'] = hinos[1] if len(hinos) > 1 else ''
                    
                if isinstance(detalhes_dict.get('oracoes'), list):
                    oracoes = detalhes_dict['oracoes']
                    detalhes_dict['oracao_abertura'] = oracoes[0] if len(oracoes) > 0 else ''
                    detalhes_dict['oracao_encerramento'] = oracoes[1] if len(oracoes) > 1 else ''
                    
                detalhes = detalhes_dict
            else:
                detalhes = {}
        else: # Tipo batismo
            cursor.execute("SELECT * FROM batismo WHERE ata_id=?", (ata_id,))
            detalhes = cursor.fetchone()
            if detalhes:
                detalhes_dict = dict(detalhes)
                if detalhes_dict.get('batizados'):
                    try:
                        detalhes_dict['batizados'] = json.loads(detalhes_dict['batizados'])
                    except:
                        detalhes_dict['batizados'] = []
                detalhes = detalhes_dict
            else:
                detalhes = {}
        
        # 4. Converter para PDF (template é vazio/None, resultando em "Sem Textos")
        buffer, filename, mimetype = exportar_pdf_bytes(ata, detalhes, template, filename=f"ata_simples_{ata_id}.pdf")
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype=mimetype)
        
    except Exception as e:
        print(f"======== ERRO CRÍTICO NA EXPORTAÇÃO DE PDF SIMPLES: {e} ========")
        flash(f"Erro ao exportar PDF Simples: {str(e)}", "error")
        return redirect(url_for("visualizar_ata", ata_id=ata_id))
    finally:
        conn.close()

# Rota para exportar ata sacramental como PDF formatado
@app.route("/ata/exportar_sacramental/<int:ata_id>")
@login_required
def exportar_sacramental_pdf(ata_id):
    from functions.pdf_exporters import exportar_sacramental_bytes
    conn = get_db()
    try:
        # Renderizar HTML
        ata = conn.execute(
            "SELECT * FROM atas WHERE id=? AND ala_id=?", 
            (ata_id, session['user_id'])
        ).fetchone()
        
        if not ata:
            raise ValueError("Ata não encontrada")
        
        ata = dict(ata)

        # Incluir o nome da ala (unidade) para exibição no cabeçalho do PDF
        unidade_row = conn.execute("SELECT nome FROM unidades WHERE ala_id = ? LIMIT 1", (ata.get('ala_id'),)).fetchone()
        if unidade_row:
            ata['ala_nome'] = unidade_row['nome']
        
        if ata["tipo"] != "sacramental":
            raise ValueError("Esta ata não é sacramental")
        
        detalhes = conn.execute("SELECT * FROM sacramental WHERE ata_id=?", (ata_id,)).fetchone()
        if detalhes:
            detalhes = dict(detalhes)
            if detalhes.get('hinos'):
                try:
                    hinos = json.loads(detalhes['hinos'])
                    detalhes['hino_abertura'] = hinos[0] if len(hinos) > 0 else ''
                    detalhes['hino_encerramento'] = hinos[1] if len(hinos) > 1 else ''
                except:
                    pass
            if detalhes.get('oracoes'):
                try:
                    oracoes = json.loads(detalhes['oracoes'])
                    detalhes['oracao_abertura'] = oracoes[0] if len(oracoes) > 0 else ''
                    detalhes['oracao_encerramento'] = oracoes[1] if len(oracoes) > 1 else ''
                except:
                    pass
            if detalhes.get('discursantes'):
                try:
                    detalhes['discursantes'] = json.loads(detalhes['discursantes'])
                except:
                    detalhes['discursantes'] = []
            if detalhes.get('anuncios'):
                try:
                    detalhes['anuncios'] = json.loads(detalhes['anuncios'])
                except:
                    detalhes['anuncios'] = []
        else:
            detalhes = {}
        
        # Buscar template padrão
        sem_textos = request.args.get('sem_textos', '').lower()
        template = conn.execute("SELECT * FROM templates WHERE nome = 'Sacramental Padrão'").fetchone()
        if not template:
            template = conn.execute(
                "SELECT * FROM templates WHERE tipo_template = 1"
            ).fetchone()
        if template:
            template = dict(template)
        else:
            template = {}

        # Se o usuário solicitou 'sem_textos', mantemos os títulos mas esvaziamos os textos do template
        if sem_textos in ('1', 'true', 'yes', 'on') and template:
            for k in ['boas_vindas', 'desobrigacoes', 'apoios', 'confirmacoes_batismo', 'apoio_membro_novo', 'bencao_crianca', 'sacramento', 'mensagens', 'live', 'encerramento']:
                template[k] = ''

        # Converter para PDF
        # Gerar PDF diretamente com dados (ReportLab)
        buffer, filename, mimetype = exportar_sacramental_bytes(ata, detalhes, template=template, filename=f"ata_sacramental_{ata_id}.pdf")
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype=mimetype)
    
    except Exception as e:
        print("======== ERRO CRÍTICO NA EXPORTAÇÃO DE PDF ========")
        traceback.print_exc() # Isso irá imprimir o erro detalhado no console

        flash(f"Erro ao exportar PDF: {str(e)}", "error")
        return redirect(url_for("visualizar_ata", ata_id=ata_id))
    finally:
        conn.close()

@app.template_filter('reverse_date_format')
def reverse_date_format(value):
    """Converte 'AAAA/MM/DD' para 'DD/MM/AAAA' (o template usa replace('-', '/') antes)"""
    # Se o formato original for AAAA-MM-DD, o input é AAAA/MM/DD após o replace no HTML.
    parts = value.split('/')
    if len(parts) == 3:
        # Reverte ordem: [2]DD, [1]MM, [0]AAAA
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return value

@app.route('/deletar_ata', methods=['POST'])
def deletar_ata():
    """Rota para deletar uma ata e seus detalhes relacionados."""
    # Garante que o usuário está logado
    if 'user_id' not in session:
        flash('Você precisa estar logado para realizar esta ação.', 'error')
        return redirect(url_for('login'))

    ata_id = request.form.get('ata_id', type=int)
    ala_id = session['user_id']

    if not ata_id:
        flash('ID da ata não fornecido.', 'error')
        # CORREÇÃO: O endpoint correto é 'listar_todas_atas'
        return redirect(url_for('listar_todas_atas'))

    conn = get_db()
    
    # 1. Obter o tipo da ata para saber qual tabela de detalhes deletar
    ata_info = conn.execute("SELECT tipo FROM atas WHERE id = ? AND ala_id = ?", (ata_id, ala_id)).fetchone()

    if not ata_info:
        flash('Ata não encontrada ou você não tem permissão para deletá-la.', 'error')
        conn.close()
        # CORREÇÃO: O endpoint correto é 'listar_todas_atas'
        return redirect(url_for('listar_todas_atas'))

    ata_tipo = ata_info['tipo']

    try:
        # Inicia transação
        conn.execute("BEGIN TRANSACTION")

        # 2. Deleta os detalhes relacionados (sacramental ou batismo)
        if ata_tipo == 'sacramental':
            conn.execute("DELETE FROM sacramental WHERE ata_id = ?", (ata_id,))
        elif ata_tipo == 'batismo':
            conn.execute("DELETE FROM batismo WHERE ata_id = ?", (ata_id,))
        
        # 3. Deleta a ata principal (precisa ter ala_id para segurança)
        conn.execute("DELETE FROM atas WHERE id = ? AND ala_id = ?", (ata_id, ala_id))

        # Confirma a transação
        conn.commit()
        flash(f'Ata de {ata_tipo.capitalize()} (ID: {ata_id}) deletada com sucesso!', 'success')

    except Exception as e:
        conn.rollback()
        flash(f'Erro ao deletar ata: {e}', 'error')
        
    finally:
        conn.close()

    # CORREÇÃO: O endpoint correto é 'listar_todas_atas'
    return redirect(url_for('listar_todas_atas'))

# Sistema de mensagens flash
@app.context_processor
def inject_flash_messages():
    messages = []
    return dict(flash_messages=messages)

# WebSocket para edição colaborativa em tempo real
# users_editing: room -> set of nicknames
users_editing = {}
# client_map: sid -> {'nick': <nick>, 'rooms': set([rooms])}
client_map = {}

@socketio.on('join')
def handle_join(data):
    room = data.get('ata_id')
    nick = data.get('nick', 'Anon')
    sid = request.sid

    print(f"[socket] join request room={room} nick={nick} sid={sid}")

    # normalize container
    if room not in users_editing or not isinstance(users_editing[room], set):
        users_editing[room] = set()
    users_editing[room].add(nick)

    # track which rooms this sid has joined and the nick
    info = client_map.get(sid)
    if not info:
        client_map[sid] = {'nick': nick, 'rooms': set([room])}
    else:
        client_map[sid].setdefault('rooms', set()).add(room)
        client_map[sid]['nick'] = nick

    join_room(room)
    emit('update_users', {'count': len(users_editing[room]), 'ata_id': room, 'users': list(users_editing[room])}, to=room)

    # Enviar estado atual dos campos ao cliente que acabou de entrar (ajuda na sincronização inicial)
    try:
        def _emit_room_state(rm):
            # rm pode ser 'ata-<id>' ou 'date-YYYY-MM-DD'
            conn = get_db()
            sac = None
            if rm.startswith('ata-'):
                try:
                    aid = int(rm.split('-', 1)[1])
                    sac = conn.execute('SELECT * FROM sacramental WHERE ata_id=?', (aid,)).fetchone()
                except Exception:
                    sac = None
            elif rm.startswith('date-'):
                date = rm.split('-', 1)[1]
                # Tentar usar sessão do usuário para filtrar por ala
                try:
                    if session.get('user_id'):
                        ata_row = conn.execute("SELECT * FROM atas WHERE data=? AND tipo='sacramental' AND ala_id=?", (date, session.get('user_id'))).fetchone()
                    else:
                        ata_row = conn.execute("SELECT * FROM atas WHERE data=? AND tipo='sacramental' LIMIT 1", (date,)).fetchone()
                    if ata_row:
                        sac = conn.execute('SELECT * FROM sacramental WHERE ata_id=?', (ata_row['id'],)).fetchone()
                except Exception:
                    sac = None
            if not sac:
                conn.close()
                return

            # Normalizar filas
            try:
                d1 = sac.get('discursante_1') or ''
                d2 = sac.get('discursante_2') or ''
                d3 = sac.get('ultimo_discursante') or ''
                tema = sac.get('tema') or ''
                tema_1 = sac.get('tema_1') or ''
                tema_2 = sac.get('tema_2') or ''
                tema_3 = sac.get('tema_ultimo') or ''
                obs_1 = sac.get('obs_1') or ''
                obs_2 = sac.get('obs_2') or ''
                obs_3 = sac.get('obs_ultimo') or ''
            except Exception:
                conn.close()
                return

            updates = {
                'discursante_1': d1,
                'discursante_2': d2,
                'discursante_3': d3,
                'tema': tema,
                'tema_1': tema_1,
                'tema_2': tema_2,
                'tema_3': tema_3,
                'obs_1': obs_1,
                'obs_2': obs_2,
                'obs_3': obs_3
            }

            date_payload = None
            if rm.startswith('date-'):
                date_payload = rm.split('-',1)[1]
            for name, val in updates.items():
                payload = {'ata_id': rm, 'name': name, 'value': val}
                if date_payload:
                    payload['date'] = date_payload
                print(f"[socket][emit] (join_init) to={rm} payload_name={name} value={val!r} date={date_payload}")
                socketio.emit('field_update', payload, to=rm)
            conn.close()

        _emit_room_state(room)
    except Exception as e:
        print(f"[socket] erro ao enviar estado inicial para room={room}: {e}")

@socketio.on('leave')
def handle_leave(data):
    room = data.get('ata_id')
    nick = data.get('nick')
    sid = request.sid

    print(f"[socket] leave request room={room} nick={nick} sid={sid}")

    if room in users_editing and nick in users_editing[room]:
        users_editing[room].remove(nick)
        if not users_editing[room]:
            del users_editing[room]

    if sid in client_map:
        client_map[sid].get('rooms', set()).discard(room)
        if not client_map[sid].get('rooms'):
            del client_map[sid]

    leave_room(room)
    emit('update_users', {'count': len(users_editing.get(room, [])), 'ata_id': room, 'users': list(users_editing.get(room, []))}, to=room)

@socketio.on('disconnect')
def handle_disconnect():
    sid = request.sid
    info = client_map.get(sid)
    if not info:
        return
    nick = info.get('nick')
    rooms = list(info.get('rooms', []))
    for room in rooms:
        if room in users_editing and nick in users_editing[room]:
            users_editing[room].remove(nick)
            if not users_editing[room]:
                del users_editing[room]
            emit('update_users', {'count': len(users_editing.get(room, [])), 'ata_id': room, 'users': list(users_editing.get(room, []))}, to=room)
    del client_map[sid]

@socketio.on('field_update')
def handle_field_update(data):
    room = data.get('ata_id')
    name = data.get('name')
    value = data.get('value')
    print(f"[socket] field_update room={room} name={name} value={str(value)[:60]}")
    # Include room identifier so clients can safely ignore updates for other cards
    emit('field_update', {'ata_id': room, 'name': name, 'value': value}, to=room, include_self=False)

# API endpoint to support polling client (returns current state by date or ata id)
@app.route('/api/discursantes_state')
@login_required
def api_discursantes_state():
    date = request.args.get('date')
    ata_id = request.args.get('ata_id', type=int)
    conn = get_db()
    sac = None
    ata_out = None
    if ata_id:
        sac = conn.execute('SELECT * FROM sacramental WHERE ata_id=?', (ata_id,)).fetchone()
        ata_out = ata_id
    elif date:
        ata_row = conn.execute("SELECT * FROM atas WHERE data = ? AND tipo = 'sacramental' AND ala_id = ?", (date, session['user_id'])).fetchone()
        if ata_row:
            sac = conn.execute('SELECT * FROM sacramental WHERE ata_id = ?', (ata_row['id'],)).fetchone()
            ata_out = ata_row['id']

    if not sac:
        conn.close()
        return jsonify({})

    # Converter sqlite row para dict para uso seguro de .get()
    try:
        sac = dict(sac)
    except Exception:
        pass

    try:
        d1 = sac.get('discursante_1') or ''
        d2 = sac.get('discursante_2') or ''
        d3 = sac.get('ultimo_discursante') or ''
        tema = sac.get('tema') or ''
        tema_1 = sac.get('tema_1') or ''
        tema_2 = sac.get('tema_2') or ''
        tema_3 = sac.get('tema_ultimo') or ''
        obs_1 = sac.get('obs_1') or ''
        obs_2 = sac.get('obs_2') or ''
        obs_3 = sac.get('obs_ultimo') or ''
    except Exception:
        conn.close()
        return jsonify({})

    conn.close()
    return jsonify({
        'ata_id': ata_out,
        'discursante_1': d1,
        'discursante_2': d2,
        'discursante_3': d3,
        'tema': tema,
        'tema_1': tema_1,
        'tema_2': tema_2,
        'tema_3': tema_3,
        'obs_1': obs_1,
        'obs_2': obs_2,
        'obs_3': obs_3,
        'date': date
    })

# Rota para renderizar HTML puro da ata (para conversão a PDF)
@app.route("/ata/render_html/<int:ata_id>")
@login_required
def render_ata_html(ata_id):
    """Renderiza o HTML puro (sem base.html) para conversão a PDF"""
    conn = get_db()
    ata = conn.execute(
        "SELECT * FROM atas WHERE id=? AND ala_id=?", 
        (ata_id, session['user_id'])
    ).fetchone()
    
    if not ata:
        flash("Ata não encontrada ou você não tem permissão para acessá-la.", "error")
        return redirect(url_for("index"))
        
    # Buscar template padrão
    template = None
    if ata["tipo"] == "sacramental":
        template = conn.execute(
            "SELECT * FROM templates WHERE nome = 'Sacramental Padrão'"
        ).fetchone()
        
        if not template:
            template = conn.execute(
                "SELECT * FROM templates WHERE tipo_template = 1"
            ).fetchone()
        
        if template:
            template = dict(template)
    
    # Buscar detalhes
    if ata["tipo"] == "sacramental":
        detalhes = conn.execute("SELECT * FROM sacramental WHERE ata_id=?", (ata_id,)).fetchone()
        if detalhes:
            detalhes_dict = dict(detalhes)
            if detalhes_dict.get('hinos'):
                try:
                    hinos = json.loads(detalhes_dict['hinos'])
                    detalhes_dict['hino_abertura'] = hinos[0] if len(hinos) > 0 else ''
                    detalhes_dict['hino_encerramento'] = hinos[1] if len(hinos) > 1 else ''
                except:
                    detalhes_dict['hino_abertura'] = ''
                    detalhes_dict['hino_encerramento'] = ''
                    
            if detalhes_dict.get('oracoes'):
                try:
                    oracoes = json.loads(detalhes_dict['oracoes'])
                    detalhes_dict['oracao_abertura'] = oracoes[0] if len(oracoes) > 0 else ''
                    detalhes_dict['oracao_encerramento'] = oracoes[1] if len(oracoes) > 1 else ''
                except:
                    detalhes_dict['oracao_abertura'] = ''
                    detalhes_dict['oracao_encerramento'] = ''
                    
            if detalhes_dict.get('discursantes'):
                try:
                    detalhes_dict['discursantes'] = json.loads(detalhes_dict['discursantes'])
                except:
                    detalhes_dict['discursantes'] = []
                    
            if detalhes_dict.get('anuncios'):
                try:
                    detalhes_dict['anuncios'] = json.loads(detalhes_dict['anuncios'])
                except:
                    detalhes_dict['anuncios'] = []
                    
            detalhes = detalhes_dict
        else:
            detalhes = {}
    else:
        detalhes = conn.execute("SELECT * FROM batismo WHERE ata_id=?", (ata_id,)).fetchone()
        if detalhes:
            detalhes_dict = dict(detalhes)
            if detalhes_dict.get('batizados'):
                try:
                    detalhes_dict['batizados'] = json.loads(detalhes_dict['batizados'])
                except:
                    detalhes_dict['batizados'] = []
            detalhes = detalhes_dict
        else:
            detalhes = {}
    
    conn.close()
    
    # Renderizar template SEM base.html (use um template dedicado ou renderize inline)
    return render_template("visualizar_ata_pdf.html", ata=ata, detalhes=detalhes, template=template)

@app.after_request
def add_header(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    return response

# Rodar o app
if __name__ == "__main__":
    # 1. Inicializa o banco de dados antes de subir o servidor
    init_db()
    
    # 2. Pega as configurações de ambiente
    port = int(os.environ.get('PORT', 5000))
    # Se estiver no seu PC, força debug=True. No servidor (Render), usa a variável de ambiente.
    debug_mode = os.environ.get('DEBUG', 'True').lower() == 'true'
    
    # 3. Roda APENAS o socketio.run (ele já gerencia o app do Flask por baixo)
    socketio.run(app, 
                 host='0.0.0.0', 
                 port=port, 
                 debug=debug_mode,
                 allow_unsafe_werkzeug=True)